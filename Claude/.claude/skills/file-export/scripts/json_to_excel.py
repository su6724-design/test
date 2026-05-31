"""analysis.json을 엑셀 파일(.xlsx)로 변환 — 요약 시트 + 출처 목록 시트"""
import sys
import json
import os
from datetime import datetime


def convert(json_path: str, xlsx_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl이 설치되지 않았습니다. 설치: pip install openpyxl")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # ── 색상 정의 ──
    BLUE_HEADER = "2563EB"
    LIGHT_BLUE  = "EFF6FF"
    WHITE       = "FFFFFF"
    IMPORTANCE_COLORS = {"상": "FEE2E2", "중": "FEF3C7", "하": "F0FDF4"}

    def header_style(cell, text):
        cell.value = text
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=BLUE_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def thin_border():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    # ════════════════════════════════
    #  시트 1: 요약
    # ════════════════════════════════
    ws1 = wb.active
    ws1.title = "요약"

    rows = [
        ("키워드",   data.get("keyword", "")),
        ("생성일",   data.get("analyzed_at", datetime.now().isoformat())[:10]),
        ("핵심 요약", data.get("summary", "")),
        ("포인트 수", str(len(data.get("key_points", [])))),
    ]

    for r, (label, value) in enumerate(rows, start=2):
        label_cell = ws1.cell(row=r, column=2, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        label_cell.alignment = Alignment(vertical="center")
        label_cell.border = thin_border()

        value_cell = ws1.cell(row=r, column=3, value=value)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        value_cell.border = thin_border()

    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 60
    ws1.row_dimensions[4].height = 80  # 핵심 요약 행 높이

    # 핵심 포인트 목록 (시트 1 하단)
    if data.get("key_points"):
        ws1.cell(row=7, column=2, value="핵심 포인트 목록").font = Font(bold=True, size=12)

        headers = ["번호", "포인트 내용", "중요도"]
        for c, h in enumerate(headers, start=2):
            header_style(ws1.cell(row=8, column=c), h)

        for i, pt in enumerate(data["key_points"], start=1):
            row = 8 + i
            ws1.cell(row=row, column=2, value=i).alignment = Alignment(horizontal="center")
            content_cell = ws1.cell(row=row, column=3, value=pt.get("point", ""))
            content_cell.alignment = Alignment(wrap_text=True, vertical="top")
            imp = pt.get("importance", "중")
            imp_cell = ws1.cell(row=row, column=4, value=imp)
            imp_cell.alignment = Alignment(horizontal="center")
            imp_cell.fill = PatternFill("solid", fgColor=IMPORTANCE_COLORS.get(imp, WHITE))
            for c in [2, 3, 4]:
                ws1.cell(row=row, column=c).border = thin_border()
            ws1.row_dimensions[row].height = 40

        ws1.column_dimensions["D"].width = 10

    # ════════════════════════════════
    #  시트 2: 출처 목록
    # ════════════════════════════════
    ws2 = wb.create_sheet("출처 목록")

    headers2 = ["번호", "제목", "URL", "날짜", "중요도"]
    col_widths = [6, 45, 50, 14, 10]
    for c, (h, w) in enumerate(zip(headers2, col_widths), start=1):
        header_style(ws2.cell(row=1, column=c), h)
        ws2.column_dimensions[get_column_letter(c)].width = w

    for i, pt in enumerate(data.get("key_points", []), start=1):
        row = 1 + i
        ws2.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=2, value=pt.get("point", "")[:80]).alignment = Alignment(wrap_text=True)
        url = pt.get("source_url", "")
        url_cell = ws2.cell(row=row, column=3, value=url)
        url_cell.font = Font(color="2563EB", underline="single")
        url_cell.alignment = Alignment(wrap_text=True)
        ws2.cell(row=row, column=4, value="")  # 날짜 (raw_data에서 가져올 수 있으면 추가)
        imp = pt.get("importance", "중")
        imp_cell = ws2.cell(row=row, column=5, value=imp)
        imp_cell.alignment = Alignment(horizontal="center")
        imp_cell.fill = PatternFill("solid", fgColor=IMPORTANCE_COLORS.get(imp, WHITE))
        for c in range(1, 6):
            ws2.cell(row=row, column=c).border = thin_border()
        ws2.row_dimensions[row].height = 36

    ws2.freeze_panes = "A2"

    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb.save(xlsx_path)
    print(f"엑셀 저장 완료: {xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("사용법: python json_to_excel.py <analysis.json> <출력.xlsx>")
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
