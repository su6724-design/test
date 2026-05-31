import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from openpyxl import load_workbook
from datetime import datetime
import calendar
import threading
import pandas as pd
from pathlib import Path
from openpyxl.styles import Alignment, Font


def ftp_list_001(ftp_folder: str, setting_dt: datetime, st2_modified_dt: datetime, out_filename="FTP_List.csv"):
    ftp_path = Path(ftp_folder)
    if not ftp_path.is_dir():
        raise ValueError(f"FTP 폴더가 올바르지 않습니다: {ftp_folder}")

    out_path = ftp_path.parent / out_filename
    csv_files = sorted(ftp_path.glob("*.csv"))  # 필요하면 rglob로 변경

    if not csv_files:
        raise FileNotFoundError(f"FTP 폴더 안에 CSV 파일이 없습니다: {ftp_folder}")

    dfs = []
    base_cols = None

    for fp in csv_files:
        df = pd.read_csv(fp, encoding="utf-8-sig", dtype=str)

        if base_cols is None:
            base_cols = list(df.columns)

        extra_cols = [c for c in df.columns if c not in base_cols]
        df = df[base_cols + extra_cols]

        df.insert(0, "_source_file", fp.name)
        dfs.append(df)

    merged = pd.concat(dfs, ignore_index=True)

    # =========================
    # ✅ B열 날짜 파싱 + ST2 수정시간 이후 행 삭제
    # =========================
    cols = list(merged.columns)
    if len(cols) < 4:
        raise ValueError("컬럼 수가 4개 미만이라 B/C/D 기준 작업이 불가능합니다.")

    b_col = cols[1]              # 엑셀 B열
    cd_cols = [cols[2], cols[3]] # 엑셀 C,D열

    # B열: '2026-01-09 10:43:20 AM' -> datetime 변환
    merged["_B_DT_"] = pd.to_datetime(
        merged[b_col].astype(str).str.replace(r"\s+", " ", regex=True).str.strip(),
        errors="coerce"
    )

    # 파싱 실패 행 제거 (원하면 유지로 바꿀 수 있음)
    merged = merged[merged["_B_DT_"].notna()].copy()

    # ✅ ST2 수정시간 이후 데이터 삭제 (B열 시간이 st2_modified_dt 보다 크면 제거)
    merged = merged[
        (merged["_B_DT_"] >= setting_dt) &
        (merged["_B_DT_"] <= st2_modified_dt)
    ].copy()

    # =========================
    # ✅ B 내림차순 → C,D 중복제거(큰 B 유지) → B 오름차순
    # =========================
    merged = merged.sort_values(by="_B_DT_", ascending=False)

    merged = merged.drop_duplicates(subset=cd_cols, keep="first")

    merged = merged.sort_values(by="_B_DT_", ascending=True).reset_index(drop=True)

    # 임시 컬럼 제거
    merged.drop(columns=["_B_DT_"], inplace=True)

    # 저장
    merged.to_csv(out_path, index=False, encoding="utf-8-sig")

    FTP_List = merged.to_dict(orient="records")
    return FTP_List, str(out_path), [str(p) for p in csv_files]




def xy_po_001(print_excel_path: str, FTP_List: list[dict], spec_x: float, spec_y: float):
    wb = load_workbook(print_excel_path)

    src_sheet_name = "1300-1200_Repair"
    new_sheet_name = "XY_PO"

    if src_sheet_name not in wb.sheetnames:
        raise ValueError(f"'{src_sheet_name}' 시트가 없습니다.")

    ws_src = wb[src_sheet_name]

    # 기존 XY_PO 있으면 삭제
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    # 시트 복제
    ws_new = wb.copy_worksheet(ws_src)
    ws_new.title = new_sheet_name

    # =========================
    # 1) C열=4만 남기기 (헤더 1~2행 유지, 데이터는 3행부터)
    # =========================
    rows_to_delete = []
    last_row = ws_new.max_row  # delete_rows 전에 고정
    for row in range(3, last_row + 1):
        val = ws_new.cell(row=row, column=3).value  # (XY_PO의) C열
        try:
            if int(float(str(val).strip())) != 4:
                rows_to_delete.append(row)
        except:
            rows_to_delete.append(row)

    for r in reversed(rows_to_delete):
        ws_new.delete_rows(r)

    # =========================
    # 2) 매칭 후 값 채우기
    # =========================
    if not FTP_List:
        wb.save(print_excel_path)
        return

    ftp_cols = list(FTP_List[0].keys())
    if len(ftp_cols) < 8:
        raise ValueError("FTP_List 컬럼이 8개 미만이라 C,D,G,H를 사용할 수 없습니다.")

    ftp_C = ftp_cols[2]
    ftp_D = ftp_cols[3]
    ftp_G = ftp_cols[6]
    ftp_H = ftp_cols[7]

    def norm(v):
        if v is None:
            return ""
        return str(v).strip()

    def to_number(v):
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        try:
            f = float(s)
            if f.is_integer():
                return int(f)
            return f
        except:
            return v  # 숫자 아니면 그대로(텍스트)

    # (FTP C, FTP D) -> (C,D,G,H) 매핑
    ftp_map = {}
    for rec in FTP_List:
        key = (norm(rec.get(ftp_C)), norm(rec.get(ftp_D)))
        ftp_map[key] = (
            rec.get(ftp_C),
            rec.get(ftp_D),
            rec.get(ftp_G),
            rec.get(ftp_H),
        )

    # XY_PO: 3행부터 데이터, 키는 A(1),B(2)
    for row in range(3, ws_new.max_row + 1):
        xy_a = ws_new.cell(row=row, column=1).value  # A열
        xy_b = ws_new.cell(row=row, column=2).value  # B열
        key = (norm(xy_a), norm(xy_b))

        if key in ftp_map:
            c_val, d_val, g_val, h_val = ftp_map[key]
            ws_new.cell(row=row, column=7).value  = to_number(c_val)  # G <- FTP C
            ws_new.cell(row=row, column=8).value  = to_number(d_val)  # H <- FTP D
            ws_new.cell(row=row, column=9).value  = to_number(g_val)  # I <- FTP G
            ws_new.cell(row=row, column=10).value = to_number(h_val)  # J <- FTP H

    # =========================
    # 3) K열에 판정 "값" 입력 (3행부터)  ← 엑셀 수식 X, 파이썬 계산 O
    # =========================
    judge_align = Alignment(horizontal="center", vertical="center")
    judge_font  = Font(bold=True)

    def to_float(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s == "":
            return None
        try:
            return float(s)
        except:
            return None

    for row in range(3, ws_new.max_row + 1):
        x = to_float(ws_new.cell(row=row, column=9).value)   # I열 (X)
        y = to_float(ws_new.cell(row=row, column=10).value)  # J열 (Y)

        if x is None or y is None:
            result = "-"
        else:
            ax, ay = abs(x), abs(y)
            if ax > spec_x and ay > spec_y:
                result = "Y"
            elif ax > spec_x:
                result = "X"
            elif ay > spec_y:
                result = "Y"
            else:
                result = "-"

        cell = ws_new.cell(row=row, column=11)  # K열
        cell.value = result
        cell.alignment = judge_align
        cell.font = judge_font

    # =========================
    # 2.5) 헤더 작성 (G2~K2) + 가운데 정렬 + Bold
    # =========================
    header_align = Alignment(horizontal="center", vertical="center")
    header_font  = Font(bold=True)

    headers = {
        7: "DUT",    # G
        8: "PAD",    # H
        9: "X",      # I
        10: "Y",     # J
        11: "판정",  # K
    }

    for col, text in headers.items():
        cell = ws_new.cell(row=2, column=col)
        cell.value = text
        cell.alignment = header_align
        cell.font = header_font

    # =========================
    # 4) PRINT 시트에 XY_PO 판정(K열) 써넣기 (확장 버전)
    # =========================
    if "PRINT" in wb.sheetnames:
        ws_print = wb["PRINT"]

        # XY_PO (A,B) -> K 매핑
        xy_to_k = {}
        for r in range(3, ws_new.max_row + 1):
            a = ws_new.cell(r, 1).value
            b = ws_new.cell(r, 2).value
            k = ws_new.cell(r, 11).value
            key = (norm(a), norm(b))
            if key != ("", ""):
                xy_to_k[key] = k

        # PRINT 처리
        for r in range(3, ws_print.max_row + 1):
            d = ws_print.cell(r, 6).value  # F열 (DUT)
            e = ws_print.cell(r, 7).value  # G열 (PAD)

            if d is None or e is None:
                continue

            d_norm = norm(d)
            e_str = str(e)

            results = []  # 찾은 결과들을 담을 리스트

            # 1️⃣ 콤마(,) 기준 분리
            for part in e_str.split(","):
                part = part.strip()

                if not part:
                    continue

                # 2️⃣ 물결(~)이 있으면 범위 전체 PAD 확인
                if "~" in part:
                    try:
                        start_p, end_p = [int(x.strip()) for x in part.split("~")]
                        for p in range(start_p, end_p + 1):
                            key = (d_norm, str(p))
                            if key in xy_to_k:
                                val = xy_to_k[key]
                                if val == "-":
                                    val = "Z"
                                results.append(str(val))
                    except:
                        pass
                    continue

                key = (d_norm, norm(part))

                if key in xy_to_k:
                    val = xy_to_k[key]
                    if val == "-":
                        val = "Z"
                    results.append(str(val))

            # 3️⃣ X/Y/Z 개수를 B/C/D열에 각각 기록
            if results:
                x_count = results.count("X")
                y_count = results.count("Y")
                z_count = results.count("Z")
                if x_count:
                    ws_print.cell(r, 2).value = x_count
                if y_count:
                    ws_print.cell(r, 3).value = y_count
                if z_count:
                    ws_print.cell(r, 4).value = z_count

    wb.save(print_excel_path)



class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("X Y Z Fail 분류")
        self.geometry("820x280")
        self.resizable(False, False)

        self._build_ui()

    def _build_ui(self):
        pad_y = 8

        tk.Label(self, text="세팅 시간 :").grid(row=0, column=0, sticky="w", padx=12, pady=pad_y)

        now = datetime.now()

        self.year_var = tk.IntVar(value=now.year)
        self.mon_var  = tk.IntVar(value=now.month)
        self.day_var  = tk.IntVar(value=now.day)
        self.hour_var = tk.IntVar(value=now.hour)
        self.min_var  = tk.IntVar(value=now.minute)

        time_frame = tk.Frame(self)
        time_frame.grid(row=0, column=1, columnspan=2, sticky="w", padx=8, pady=pad_y)

        self.year_values = list(range(2000, 2101))
        self.mon_values  = list(range(1, 13))
        self.hour_values = list(range(0, 24))
        self.min_values  = list(range(0, 60))

        self.cmb_year = self._combo(time_frame, self.year_var, self.year_values, "년", width=6)
        self.cmb_mon  = self._combo(time_frame, self.mon_var,  self.mon_values,  "월", width=4)
        self.cmb_day  = self._combo(time_frame, self.day_var,  [],               "일", width=4)
        self.cmb_hour = self._combo(time_frame, self.hour_var, self.hour_values, "시", width=4)
        self.cmb_min  = self._combo(time_frame, self.min_var,  self.min_values,  "분", width=4)

        self._refresh_days()

        self.cmb_year.bind("<<ComboboxSelected>>", lambda e: self._refresh_days())
        self.cmb_mon.bind("<<ComboboxSelected>>",  lambda e: self._refresh_days())

        tk.Label(self, text="Print 파일 :").grid(row=1, column=0, sticky="w", padx=12, pady=pad_y)
        self.print_path = tk.StringVar()
        tk.Entry(self, textvariable=self.print_path, width=92).grid(row=1, column=1, sticky="w", padx=8)
        tk.Button(self, text="찾기", width=6, command=self._pick_print).grid(row=1, column=2)

        tk.Label(self, text="ST2 파일 :").grid(row=2, column=0, sticky="w", padx=12, pady=pad_y)
        self.st2_path = tk.StringVar()
        tk.Entry(self, textvariable=self.st2_path, width=92).grid(row=2, column=1, sticky="w", padx=8)
        tk.Button(self, text="찾기", width=6, command=self._pick_st2).grid(row=2, column=2)

        tk.Label(self, text="FTP 폴더 :").grid(row=3, column=0, sticky="w", padx=12, pady=pad_y)
        self.ftp_dir = tk.StringVar()
        tk.Entry(self, textvariable=self.ftp_dir, width=92).grid(row=3, column=1, sticky="w", padx=8)
        tk.Button(self, text="찾기", width=6, command=self._pick_ftp).grid(row=3, column=2)

        # Spec X/Y 입력
        tk.Label(self, text="Spec X Y :").grid(row=4, column=0, sticky="w", padx=12, pady=pad_y)
        self.spec_x = tk.StringVar(value="3")
        self.spec_y = tk.StringVar(value="3")
        spec_frame = tk.Frame(self)
        spec_frame.grid(row=4, column=1, sticky="w", padx=8, pady=pad_y)
        tk.Label(spec_frame, text="X >").pack(side="left")
        tk.Entry(spec_frame, textvariable=self.spec_x, width=8, justify="center").pack(side="left", padx=(4, 12))
        tk.Label(spec_frame, text="Y >").pack(side="left")
        tk.Entry(spec_frame, textvariable=self.spec_y, width=8, justify="center").pack(side="left", padx=(4, 0))

        # 실행 버튼/상태 라벨은 한 줄 아래로 내림
        self.btn_run = tk.Button(self, text="실행", width=10, command=self._run)
        self.btn_run.grid(row=5, column=1, pady=18)
        self.status_var = tk.StringVar(value="")
        tk.Label(self, textvariable=self.status_var).grid(row=6, column=1, sticky="w", pady=2)

    def _combo(self, parent, var, values, suffix, width=4):
        cmb = ttk.Combobox(parent, state="readonly", width=width, justify="center")
        cmb.pack(side="left")
        tk.Label(parent, text=suffix, padx=6).pack(side="left")

        self._set_combo_values(cmb, values)
        self._set_combo_current_from_var(cmb, var)

        def on_select(_e=None):
            try:
                var.set(int(cmb.get()))
            except:
                pass
        cmb.bind("<<ComboboxSelected>>", on_select)
        return cmb

    def _set_combo_values(self, cmb, values):
        cmb["values"] = [str(v) for v in values]

    def _set_combo_current_from_var(self, cmb, var):
        target = str(var.get())
        vals = list(cmb["values"])
        if target in vals:
            cmb.set(target)
        elif vals:
            cmb.set(vals[0])

    def _refresh_days(self):
        y = int(self.year_var.get())
        m = int(self.mon_var.get())
        last_day = calendar.monthrange(y, m)[1]
        day_values = list(range(1, last_day + 1))

        self._set_combo_values(self.cmb_day, day_values)

        cur_day = int(self.day_var.get())
        if cur_day > last_day:
            self.day_var.set(last_day)

        self._set_combo_current_from_var(self.cmb_day, self.day_var)

    # =========================
    # Pickers
    # =========================
    def _pick_print(self):
        path = filedialog.askopenfilename(
            title="Print 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx;*.xlsm;*.xls"), ("모든 파일", "*.*")]
        )
        if path:
            self.print_path.set(path)

    def _pick_st2(self):
        path = filedialog.askopenfilename(
            title="ST2 파일 선택",
            filetypes=[("ST2 파일", "*.st2"), ("모든 파일", "*.*")]
        )
        if not path:
            return

        # ST2 경로만 저장 (세팅 시간은 변경하지 않음)
        self.st2_path.set(path)

        # ST2 수정 시간 읽기
        mtime = os.path.getmtime(path)
        dt = datetime.fromtimestamp(mtime)

        # ✅ 상태 표시줄에 마감 시간 표시
        self.status_var.set(f"마감 시간: {dt:%Y-%m-%d %H:%M:%S}")

        # day 콤보 범위만 보정 (세팅 시간 값은 그대로)
        self._refresh_days()

    def _pick_ftp(self):
        path = filedialog.askdirectory(title="FTP 폴더 선택")
        if path:
            self.ftp_dir.set(path)

    # =========================
    # Run
    # =========================
    def _run(self):
        # 중복 실행 방지
        self.btn_run.config(state="disabled")
        self.status_var.set("작업 중...")

        t = threading.Thread(target=self._run_worker, daemon=True)
        t.start()

    def _run_worker(self):
        try:
            # 1) 세팅 시간 만들기
            try:
                setting_dt = datetime(
                    int(self.cmb_year.get()),
                    int(self.cmb_mon.get()),
                    int(self.cmb_day.get()),
                    int(self.cmb_hour.get()),
                    int(self.cmb_min.get())
                )
            except Exception:
                self.after(0, lambda: messagebox.showerror("오류", "세팅 시간이 올바르지 않습니다."))
                return

            # 2) 입력값 읽기
            print_file = self.print_path.get().strip()
            st2_file   = self.st2_path.get().strip()
            ftp_folder = self.ftp_dir.get().strip()

            # 3) 입력값 검증
            if not print_file or not os.path.isfile(print_file):
                self.after(0, lambda: messagebox.showerror("입력 오류", "Print 파일 경로가 올바르지 않습니다."))
                return
            if not st2_file or not os.path.isfile(st2_file):
                self.after(0, lambda: messagebox.showerror("입력 오류", "ST2 파일 경로가 올바르지 않습니다."))
                return
            if not ftp_folder or not os.path.isdir(ftp_folder):
                self.after(0, lambda: messagebox.showerror("입력 오류", "FTP 폴더 경로가 올바르지 않습니다."))
                return

            # 4) ST2 수정 시간 읽기
            st2_mtime = os.path.getmtime(st2_file)
            st2_modified_dt = datetime.fromtimestamp(st2_mtime)

            # 5) FTP CSV 병합/필터링
            try:
                FTP_List, saved_path, merged_files = ftp_list_001(
                    ftp_folder=ftp_folder,
                    setting_dt=setting_dt,
                    st2_modified_dt=st2_modified_dt
                )
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("병합 오류", str(e)))
                return

            # =========================
            # _run_worker()에서 Spec_X/Y 읽어서 xy_po_001에 전달
            # =========================
            try:
                spec_x = float(self.spec_x.get().strip())
                spec_y = float(self.spec_y.get().strip())
            except Exception:
                self.after(0, lambda: messagebox.showerror("입력 오류", "Spec X/Y 값이 올바르지 않습니다."))
                return

            # 6) Print 파일에 XY_PO 시트 생성 + C열=4만 남기기
            try:
                xy_po_001(print_file, FTP_List, spec_x, spec_y)
            except Exception as e:
                msg = str(e)
                self.after(0, lambda m=msg: messagebox.showerror("엑셀 처리 오류 or 엑셀 열림", m))
                return

            # 7) 완료 안내 (UI 스레드에서)
            def show_done():
                messagebox.showinfo(
                    "완료",
                    f"세팅 시간: {setting_dt:%Y-%m-%d %H:%M}\n"
                    f"마감 시간: {st2_modified_dt:%Y-%m-%d %H:%M}\n"
                    f"병합 파일 수: {len(merged_files)}\n"
                    f"총 행 수: {len(FTP_List)}\n"
                    f"저장 위치: {saved_path}"
                )

            self.after(0, show_done)

        except Exception as e:
            self.after(0, lambda: messagebox.showerror("오류", str(e)))

        finally:
            # 버튼/상태 복구 (UI 스레드)
            self.after(0, lambda: self.btn_run.config(state="normal"))
            self.after(0, lambda: self.status_var.set("완료"))



if __name__ == "__main__":
    App().mainloop()
