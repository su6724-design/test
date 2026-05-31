import os
import sys
import shutil
import stat
import traceback
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# ------------------------
# 전역 변수
# ------------------------
st2_file_path = ""
csv_file_path = ""
file_1200_path = ""
file_1300_path = ""
sl_values = []      #260127 이후 추가

# ------------------------
# GUI 선택 함수
# ------------------------
def parse_sl_input(sl_input):
    values = []
    for v in sl_input.split(','):
        val = v.strip().lower()
        if val == 'l':
            values.append('Long')
        elif val == 's':
            values.append('Short')
        elif val:
            values.append(v.strip())
    return values

def select_st2_file():
    path = filedialog.askopenfilename(
        title="st2 파일 선택",
        filetypes=[("ST2 파일", "*.st2"), ("모든 파일", "*.*")]
    )
    if path:
        st2_entry.delete(0, tk.END)
        st2_entry.insert(0, path)

def select_csv_file():
    path = filedialog.askopenfilename(
        title="CSV 파일 선택",
        filetypes=[("CSV 파일", "*.csv"), ("모든 파일", "*.*")]
    )
    if path:
        csv_entry.delete(0, tk.END)
        csv_entry.insert(0, path)

def select_repair_1200():
    path = filedialog.askopenfilename(
        title="1200 Repair_sheet 선택",
        filetypes=[("Excel 파일", "*.xlsx;*.xls"), ("모든 파일", "*.*")]
    )
    if path:
        repair1200_entry.delete(0, tk.END)
        repair1200_entry.insert(0, path)

def select_repair_1300():
    path = filedialog.askopenfilename(
        title="1300 Repair_sheet 선택",
        filetypes=[("Excel 파일", "*.xlsx;*.xls"), ("모든 파일", "*.*")]
    )
    if path:
        repair1300_entry.delete(0, tk.END)
        repair1300_entry.insert(0, path)
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext in ['.xlsx', '.xls']:
                df_tmp = pd.read_excel(path, header=None)
            else:
                try:
                    df_tmp = pd.read_csv(path, encoding='cp949', header=None)
                except UnicodeDecodeError:
                    df_tmp = pd.read_csv(path, encoding='utf-8', header=None)

            def is_numeric(v):
                try:
                    float(v)
                    return True
                except (ValueError, TypeError):
                    return False

            num_cols = df_tmp.shape[1]
            lbl_col_count.config(text=f"{num_cols}열")
            if num_cols == 7:
                raw_vals = df_tmp.iloc[:, 4].dropna().astype(str).str.strip()
                unique_vals = sorted(set(
                    v for v in raw_vals
                    if v and v.lower() != 's/l' and not is_numeric(v)
                ))
                entry_sl['values'] = unique_vals
                entry_sl.set(unique_vals[0] if unique_vals else 'S/L')
            elif num_cols == 6:
                entry_sl['values'] = ['S/L']
                entry_sl.set('S/L')
            else:
                messagebox.showwarning("열 수 오류",
                    f"1300 파일이 {num_cols}열입니다.\n7열로 수정해주세요.")
                entry_sl['values'] = []
                entry_sl.set('')
        except Exception as e:
            print(f"1300 파일 E열 읽기 오류: {e}")

# ------------------------
# 선택 완료 → main 실행
# ------------------------
def run_main():
    global st2_file_path, csv_file_path, file_1200_path, file_1300_path, sl_values

    st2_file_path = st2_entry.get()
    csv_file_path = csv_entry.get()
    file_1200_path = repair1200_entry.get()
    file_1300_path = repair1300_entry.get()
    sl_input = entry_sl.get().strip()
    sl_values = parse_sl_input(sl_input)

    print("S/L 값:", sl_values)
    print("ST2 파일:", st2_file_path)
    print("CSV 파일:", csv_file_path)
    print("Repair 1200 파일:", file_1200_path)
    print("Repair 1300 파일:", file_1300_path)


    root.destroy()  # GUI 닫고 main 실행

# ------------------------
# GUI 윈도우
# ------------------------
root = tk.Tk()
root.title("Fail_Pin or MAP")

# S/L 입력 (1300 파일 선택 시 E열 목록으로 자동 채워짐)
tk.Label(root, text="S/L 입력:").grid(row=0, column=0, padx=5, pady=5)
sl_frame = tk.Frame(root)
sl_frame.grid(row=0, column=1, padx=5, pady=5, sticky="w")
entry_sl = ttk.Combobox(sl_frame, width=25)
entry_sl.pack(side="left")
lbl_col_count = tk.Label(sl_frame, text="", anchor="w")
lbl_col_count.pack(side="left", padx=(6, 0))

# st2
tk.Label(root, text="ST2 파일:").grid(row=1, column=0, padx=5, pady=5)
st2_entry = tk.Entry(root, width=50)
st2_entry.grid(row=1, column=1, padx=5, pady=5)
tk.Button(root, text="찾기", command=select_st2_file).grid(row=1, column=2, padx=5, pady=5)

# csv
tk.Label(root, text="CSV 파일:").grid(row=2, column=0, padx=5, pady=5)
csv_entry = tk.Entry(root, width=50)
csv_entry.grid(row=2, column=1, padx=5, pady=5)
tk.Button(root, text="찾기", command=select_csv_file).grid(row=2, column=2, padx=5, pady=5)

# repair 1200
tk.Label(root, text="Repair 1200:").grid(row=3, column=0, padx=5, pady=5)
repair1200_entry = tk.Entry(root, width=50)
repair1200_entry.grid(row=3, column=1, padx=5, pady=5)
tk.Button(root, text="찾기", command=select_repair_1200).grid(row=3, column=2, padx=5, pady=5)

# repair 1300
tk.Label(root, text="Repair 1300:").grid(row=4, column=0, padx=5, pady=5)
repair1300_entry = tk.Entry(root, width=50)
repair1300_entry.grid(row=4, column=1, padx=5, pady=5)
tk.Button(root, text="찾기", command=select_repair_1300).grid(row=4, column=2, padx=5, pady=5)

# 확인 버튼
tk.Button(root, text="실행", command=run_main).grid(row=5, column=1, pady=10)

# GUI 실행
root.mainloop()

# ------------------------
# run_01 ~ run_18 정의
# ------------------------
def run_01():
    if not st2_file_path.endswith('.st2'):
        print("'.st2' 확장자의 파일만 지원합니다.")
        return
    elif not os.path.exists(st2_file_path):
        print(f"파일이 존재하지 않습니다: {st2_file_path}")
        return

    try:
        # st2 → Excel 변환
        with open(st2_file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
        # 줄 끝 개행 제거 (공백 줄은 ''으로 유지)
        cleaned_lines = [line.strip() for line in lines]
        # DataFrame으로 변환
        df_st2 = pd.DataFrame(cleaned_lines)

        # Excel 파일로 저장 (기본 시트 'st2')
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            df_st2.to_excel(writer, index=False, header=False, sheet_name='st2')

        print(f"Excel 파일로 저장 완료: {output_excel_path}")

        if not csv_file_path.endswith('.csv'):
            print("'.csv' 확장자의 파일만 지원합니다.")
            return
        elif not os.path.exists(csv_file_path):
            print(f"CSV 파일이 존재하지 않습니다: {csv_file_path}")
            return

        # CSV 불러오기 (헤더 없이 전체 데이터를 데이터로 처리)
        df_csv = pd.read_csv(csv_file_path, header=None)
        # 기존 Excel 파일에 시트 추가
        with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_csv.to_excel(writer, index=False, header=False, sheet_name='CSV')

        print("CSV 데이터를 'CSV' 시트로 추가 완료.")

    except Exception as e:
        print(f"오류 발생: {e}")



def run_02():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        source_sheet_name = 'CSV' # 원본 데이터가 있는 시트 이름

        wb = load_workbook(output_excel_path)
        if source_sheet_name not in wb.sheetnames:
            print(f"시트 '{source_sheet_name}'이(가) 존재하지 않습니다.")
            return

        ws_source = wb[source_sheet_name]

        # 'pin coodination'이 포함된 A열 셀 찾기
        target_row = None
        for row in ws_source.iter_rows(min_col=1, max_col=1):
            val = row[0].value
            if isinstance(val, str) and "pin coodination" in val.lower():
                target_row = row[0].row
                break

        if target_row is None:
            print("'pin coodination'이 포함된 셀을 찾을 수 없습니다.")
            return

        # 새 시트 생성 또는 기존 시트 선택
        new_sheet_name = 'pin coodination'
        if new_sheet_name in wb.sheetnames:
            ws_target = wb[new_sheet_name]
            # 기존 시트의 내용을 비웁니다.
            # 이 부분을 주석 처리하거나 제거해야 run_03에서 기존 데이터 위에 덮어쓰지 않고 활용할 수 있습니다.
            # run_03에서 DataFrame으로 불러올 것이므로 여기서는 굳이 비울 필요가 없습니다.
            # for row in ws_target.iter_rows():
            #     for cell in row:
            #         cell.value = None
            pass # 기존 시트가 있다면 비우는 대신 pass
        else:
            ws_target = wb.create_sheet(new_sheet_name)

        # 'pin coodination'이 포함된 행부터 전체 데이터를 복사 (openpyxl로 복사)
        # run_03에서 pandas로 다시 불러올 것이므로, 여기서 복사하는 것은 중복될 수 있습니다.
        # 그러나 이전 로직과의 일관성을 위해 유지합니다.
        for r_idx, row in enumerate(ws_source.iter_rows(min_row=target_row), 1):
            for c_idx, cell in enumerate(row, 1):
                ws_target.cell(row=r_idx, column=c_idx, value=cell.value)
        
        # 헤더 작성 (G~L)
        headers = ["Type", "PAD 거리", "정역구분", "거리함수", "끝 번호", "블럭"]
        bold_font = Font(bold=True)
        # 'pin coodination' 문자열이 있는 행 (새 시트의 1행)에 헤더를 추가합니다.
        header_row_in_new_sheet = 1 # 헤더가 들어갈 행
        for col, header in enumerate(headers, start=7):  # G열 = 7
            cell = ws_target.cell(row=header_row_in_new_sheet, column=col, value=header)
            cell.font = bold_font

        wb.save(output_excel_path)
        print(f"'{new_sheet_name}' 시트에 헤더 적용 완료.")

    except Exception as e:
        print(f"오류 발생: {e}")


def run_03():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        sheet_to_load = 'pin coodination'

        # 'pin coodination' 시트 불러오기 (헤더가 있는 첫 행부터 데이터로 간주)
        # Excel 시트의 D열과 E열이 Pandas에서는 각각 3번, 4번 컬럼 (0-indexed)에 해당합니다.
        # 헤더가 첫 행(1행)에 있으므로, 실제 데이터는 2행부터 시작하며, 헤더를 명시적으로 처리하지 않습니다.
        # 'pin coodination' 시트의 1행에는 "pin coodination ..." 문자열과 G~L 헤더가 있고,
        # 실제 계산 대상 데이터는 2행부터 시작합니다.
        # 따라서 skiprows=1로 설정하여 첫 번째 행(헤더 행)을 건너뛰고 데이터를 읽어옵니다.
        df_pin_coord = pd.read_excel(output_excel_path, sheet_name=sheet_to_load, header=None, skiprows=1)

        # H열 (PAD 거리) 계산: D(현재행) - D(현재행+1)
        # df_pin_coord[3]는 D열에 해당합니다.
        # .shift(-1)은 데이터를 한 칸 위로 올리므로, 현재 행의 값에서 다음 행의 값을 뺀다.
        df_pin_coord['H_PAD_거리'] = df_pin_coord[3] - df_pin_coord[3].shift(-1)
        # 마지막 행은 다음 값이 없으므로 NaN이 됩니다. 이를 0으로 채우거나 필요에 따라 다르게 처리할 수 있습니다.
        df_pin_coord['H_PAD_거리'] = df_pin_coord['H_PAD_거리'].fillna(0).astype(int) # 정수형으로 변환

        # I열 (정역구분) 계산: E(현재행) - E(현재행+1)
        # df_pin_coord[4]는 E열에 해당합니다.
        df_pin_coord['I_정역구분'] = df_pin_coord[4] - df_pin_coord[4].shift(-1)
        df_pin_coord['I_정역구분'] = df_pin_coord['I_정역구분'].fillna(0).astype(int) # 정수형으로 변환

        # 계산된 값을 기존 Excel 파일에 다시 쓰기
        # 'pin coodination' 시트를 mode='a' (append)로 열고 if_sheet_exists='overlay'를 사용하여
        # 기존 시트의 내용을 덮어씁니다. 이때, openpyxl 엔진을 사용해야 합니다.
        
        # Openpyxl 워크북과 시트 객체 로드
        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb[sheet_to_load]

        # 데이터 시작 행 (헤더 다음 행 = 2행)부터 값을 업데이트
        # df_pin_coord의 0번 인덱스는 Excel의 2행에 해당합니다.
        # 따라서 Excel 행 번호는 df_idx + 2가 됩니다.
        for r_idx, (df_idx, row_data) in enumerate(df_pin_coord.iterrows()):
            excel_row_idx = df_idx + 2 # Excel의 실제 행 번호 (1-based)
            
            # H열 업데이트 (8번째 컬럼)
            ws.cell(row=excel_row_idx, column=8, value=row_data['H_PAD_거리'])
            # I열 업데이트 (9번째 컬럼)
            ws.cell(row=excel_row_idx, column=9, value=row_data['I_정역구분'])

        wb.save(output_excel_path)
        print(f"'{sheet_to_load}' 시트의 H열과 I열 계산 및 업데이트 완료.")

    except Exception as e:
        print(f"run_03 오류 발생: {e}")


def run_04():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        sheet_to_load = 'pin coodination'

        # 'pin coodination' 시트 불러오기 (헤더가 있는 첫 행을 건너뛰고 데이터만 로드)
        # 엑셀의 A, G, H, I열이 Pandas에서는 각각 0, 6, 7, 8번 컬럼 (0-indexed)에 해당
        df_pin_coord = pd.read_excel(output_excel_path, sheet_name=sheet_to_load, header=None, skiprows=1)

        # J열 (거리함수) 계산: =OR(ABS(H2)>800,ABS(I2)>1000)
        # H열은 df_pin_coord[7], I열은 df_pin_coord[8]
        df_pin_coord['J_거리함수'] = (df_pin_coord[7].abs() > 800) | (df_pin_coord[8].abs() > 1000)

        # K열 (끝 번호) 계산: =IF(J2=TRUE,A2,"-")
        # J열은 df_pin_coord['J_거리함수'], A열은 df_pin_coord[0]
        df_pin_coord['K_끝번호'] = df_pin_coord.apply(lambda row: row[0] if row['J_거리함수'] else "-", axis=1)

        # L열 (블럭) 계산: =IF(G1=G2,IF(K1="-",L1+1,1),1)
        # 초기값 설정 (첫 번째 데이터 행의 L열은 1)
        df_pin_coord['L_블럭'] = 1

        # L열 계산을 위한 반복 처리
        # G열은 df_pin_coord[6], K열은 df_pin_coord['K_끝번호']
        for i in range(1, len(df_pin_coord)):
            # 이전 행의 G열과 현재 행의 G열이 같고, 이전 행의 K열이 "-"인 경우
            if df_pin_coord.loc[i, 6] == df_pin_coord.loc[i-1, 6]:
                if df_pin_coord.loc[i-1, 'K_끝번호'] == "-":
                    df_pin_coord.loc[i, 'L_블럭'] = df_pin_coord.loc[i-1, 'L_블럭'] + 1
                else:
                    df_pin_coord.loc[i, 'L_블럭'] = 1
            else:
                df_pin_coord.loc[i, 'L_블럭'] = 1

        # 계산된 값을 기존 Excel 파일에 다시 쓰기
        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb[sheet_to_load]

        # 데이터 시작 행 (헤더 다음 행 = 2행)부터 값을 업데이트
        for df_idx, row_data in df_pin_coord.iterrows():
            excel_row_idx = df_idx + 2 # Excel의 실제 행 번호 (1-based)
            
            # J열 업데이트 (10번째 컬럼)
            ws.cell(row=excel_row_idx, column=10, value=row_data['J_거리함수'])
            # K열 업데이트 (11번째 컬럼)
            ws.cell(row=excel_row_idx, column=11, value=row_data['K_끝번호'])
            # L열 업데이트 (12번째 컬럼)
            ws.cell(row=excel_row_idx, column=12, value=row_data['L_블럭'])

        wb.save(output_excel_path)
        print(f"'{sheet_to_load}' 시트의 J, K, L열 계산 및 업데이트 완료.")

    except Exception as e:
        print(f"run_04 오류 발생: {e}")


# 전역 변수 선언
type_list = []
type_lists = {}
list_block = []

def run_05():
    try:
        global type_list, type_lists  # 전역 변수 사용 선언

        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'

        wb = load_workbook(output_excel_path, data_only=True)

        # "pin coodination" 시트 바로 사용
        if "pin coodination" not in wb.sheetnames:
            print("'pin coodination' 시트를 찾을 수 없습니다.")
            return

        ws = wb["pin coodination"]

        # 데이터 시작 행 찾기
        target_row = None
        for row in ws.iter_rows(min_col=1, max_col=1):
            if isinstance(row[0].value, str) and "pin coodination" in row[0].value.lower():
                target_row = row[0].row
                break

        if target_row is None:
            print("'pin coodination' 텍스트가 포함된 셀을 찾을 수 없습니다.")
            return

        data_start_row = target_row + 1
        max_row = ws.max_row

        # G열의 타입 리스트 추출
        type_values = set()
        for row in range(data_start_row, max_row + 1):
            g_val = ws.cell(row=row, column=7).value
            if g_val:
                type_values.add(str(g_val).strip())

        type_list = sorted(type_values)
        type_lists = {f"type_list_{i}": [] for i in range(len(type_list))}

        # G열 기준으로 L열 값 분류
        for row in range(data_start_row, max_row + 1):
            g_val = ws.cell(row=row, column=7).value
            l_val = ws.cell(row=row, column=12).value

            if g_val is None or l_val is None:
                continue

            g_val_str = str(g_val).strip()
            for i, type_name in enumerate(type_list):
                if g_val_str == type_name:
                    type_lists[f"type_list_{i}"].append(l_val)

        # "Type" 시트 생성
        if "Type" in wb.sheetnames:
            del wb["Type"]
        ws_type = wb.create_sheet("Type")

        # 타입 이름 작성
        for i, type_name in enumerate(type_list):
            ws_type.cell(row=1, column=i + 1, value=type_name)

        # 해당 값 작성
        for i, key in enumerate(type_lists.keys()):
            for j, val in enumerate(type_lists[key], start=2):
                ws_type.cell(row=j, column=i + 1, value=val)

        wb.save(output_excel_path)
        print("'Type' 시트 생성 및 L열 값 분류 완료.")
        print("Type List:", type_list)
        print("Data Extracted:", type_lists)

    except Exception as e:
        print(f"오류 발생: {e}")


def run_06():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'

        wb = load_workbook(output_excel_path, data_only=True)

        ws = wb.active
        ws.title = "st2"

        map_dut_sheet = wb.create_sheet(title="MAP_DUT")
        map_pad_sheet = wb.create_sheet(title="MAP_PAD")

        dut_list = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if cell.value:
                    if cell.value == '[MAP_PAD]':
                        break
                    split_values = cell.value.split('=')
                    if len(split_values) == 2:
                        dut_list.append(split_values)
                    else:
                        dut_list.append([split_values[0], ''])
            if cell.value == '[MAP_PAD]':
                break

        map_dut_sheet.append(['DUT', 'PAD'])
        for i, (d1, d2) in enumerate(dut_list, start=2):
            map_dut_sheet.cell(row=i, column=1, value=d1)
            map_dut_sheet.cell(row=i, column=2, value=d2)

        map_pad_processing = False
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if cell.value:
                    if cell.value == '[MAP_PAD]':
                        map_pad_processing = True
                        continue
                    if map_pad_processing:
                        map_pad_values = cell.value.split('_')
                        row_values = []
                        for value in map_pad_values:
                            row_values.extend(value.split('='))
                        if row_values[0] != 'DUT':
                            map_pad_sheet.append(row_values)

        map_pad_sheet.insert_rows(1)
        map_pad_sheet.cell(row=1, column=1, value='DUT')
        map_pad_sheet.cell(row=1, column=2, value='PAD')
        map_pad_sheet.cell(row=1, column=3, value='CODE')

        wb.save(output_excel_path)

    except Exception as e:
        print(f"오류 발생: {e}")


def run_07():
    from openpyxl import load_workbook

    try:
        global list_block
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path, data_only=True)

        csv_ws = wb["CSV"]
        type_ws = wb["Type"]

        # Step 1: "substrate dut coodination" 찾기
        target_row = None
        for row in csv_ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == "substrate dut coodination":
                    target_row = cell.row
                    break
            if target_row:
                break

        if not target_row:
            print("CSV 시트에서 'substrate dut coodination'을 찾을 수 없습니다.")
            return

        # Step 2: 다음 행부터 D열 값들 모두 추출
        d_values = []
        row = target_row + 1
        while True:
            cell_value = csv_ws.cell(row=row, column=4).value
            if cell_value is None:
                break
            d_values.append(cell_value)
            row += 1

        # Step 3: Type 시트 1행에서 매칭되는 열들 찾기 (여러 열 가능성 포함)
        matched_cols = []
        for d_val in d_values:
            for col in range(1, type_ws.max_column + 1):
                header = type_ws.cell(row=1, column=col).value
                if header and str(header).strip() == str(d_val).strip():
                    matched_cols.append(col)
                    break

        if not matched_cols:
            # 1) 콘솔에도 남기고
            print(f"Type 시트에서 {d_values} 중 매칭되는 열을 찾을 수 없습니다.")

            # 2) 팝업 띄우기
            try:
                root = tk.Tk()
                root.withdraw()  # 빈 창 숨김
                root.attributes("-topmost", True)  # 팝업을 맨 위로(선택)
                messagebox.showerror("TYPE 오류", "CSV 파일에서 TYPE 매칭되게 수정해주세요")
                root.destroy()
            except Exception as _:
                # 혹시 tkinter가 안 되면 콘솔만 출력되고 종료
                pass

            # 3) 즉시 종료
            sys.exit(1)

        # Step 4: 각 매칭된 열에서 아래 값들을 저장
        list_block = []  # 전역 변수 초기화
        for col in matched_cols:
            for row in range(2, type_ws.max_row + 1):
                val = type_ws.cell(row=row, column=col).value
                if val is not None:
                    list_block.append(val)

        # D열, list_block 숫자 값들의 개수 출력
        print(f"D열 개수: {len([v for v in matched_cols if isinstance(v, (int, float))])}")
        print(f"list_block 개수: {len([v for v in list_block if isinstance(v, (int, float))])}")

        # Step 5: MAP_PAD 시트에 D열(4열)에 '블럭' 헤더 추가하고 list_block 작성
        if "MAP_PAD" not in wb.sheetnames:
            print("MAP_PAD 시트를 찾을 수 없습니다.")
            return

        map_pad_ws = wb["MAP_PAD"]

        block_col = 4  # D열
        block_header_row = 1
        map_pad_ws.cell(row=block_header_row, column=block_col, value="블럭")

        for i, val in enumerate(list_block):
            map_pad_ws.cell(row=block_header_row + 1 + i, column=block_col, value=val)

        # 저장
        wb.save(output_excel_path)
        print("MAP_PAD 시트에 '블럭' 열과 list_block 데이터를 작성했습니다.")

    except Exception as e:
        print(f"run_07에서 오류 발생: {e}")


def run_08():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path)

        if "MAP_PAD" not in wb.sheetnames:
            print("MAP_PAD 시트를 찾을 수 없습니다.")
            return
        ws = wb["MAP_PAD"]

        # 데이터 추출
        data = []
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=4, values_only=True):
            data.append(row)

        df = pd.DataFrame(data, columns=["CODE", "BLOCK"])
        df["CODE"] = pd.to_numeric(df["CODE"], errors="coerce").fillna(-1).astype(int)
        df["BLOCK"] = pd.to_numeric(df["BLOCK"], errors="coerce").fillna(-1).astype(int)
        df["New_CODE"] = None

        # BLOCK 연속성 기반 그룹 생성
        df["BLOCK_DIFF"] = df["BLOCK"].diff().fillna(1)
        df["BLOCK_GROUP"] = (df["BLOCK_DIFF"] != 1).cumsum()

        # 그룹 단위 처리
        for _, group in df.groupby("BLOCK_GROUP"):
            codes = group["CODE"].tolist()
            blocks = group["BLOCK"].tolist()
            indexes = group.index.tolist()

            for i, idx in enumerate(indexes):
                code = codes[i]

                # === [3, 5, 6] 처리 ===
                if code in [3, 5, 6]:
                    if 0 in codes[:i]:  # 앞쪽에 0이 있으면
                        df.at[idx, "New_CODE"] = 4
                    else:
                        df.at[idx, "New_CODE"] = 3

                # === [4, 8, 9] 처리 ===
                elif code in [4, 8, 9]:
                    is_last = (i == len(codes) - 1)

                    # ① 바로 앞이 3/5/6이고 현재가 그룹 마지막일 경우 → 10
                    if i > 0 and codes[i - 1] in [3, 5, 6] and is_last:
                        df.at[idx, "New_CODE"] = 10
                        continue

                    # ② 앞쪽에 0이 있으면 → 4
                    if 0 in codes[:i]:
                        df.at[idx, "New_CODE"] = 4
                        continue

                    # ③ 이후 코드 순서대로 판단
                    next_codes = codes[i+1:]
                    result = 4  # 기본값
                    for nc in next_codes:
                        if nc in [3, 5, 6]:
                            result = 3
                            break
                        elif nc == 0:
                            break  # 0 먼저 나오면 그대로 4 유지
                    df.at[idx, "New_CODE"] = result

                # === 0 처리 ===
                elif code == 0:
                    found_nonzero = False
                    for j in range(i + 1, len(codes)):
                        if codes[j] != 0:
                            found_nonzero = True
                            for k in range(i, j + 1):
                                df.at[indexes[k], "New_CODE"] = 4
                            break
                    if not found_nonzero:
                        df.at[idx, "New_CODE"] = 0

        # 결과 기록
        ws.cell(row=1, column=5, value="New_CODE")
        for i, val in enumerate(df["New_CODE"], start=2):
            ws.cell(row=i, column=5, value=int(val) if pd.notna(val) else "")

        wb.save(output_excel_path)
        print("✅ run_08: 'MAP_PAD' 시트의 E열 'New_CODE' 계산 완료")

    except Exception as e:
        print(f"❌ run_08에서 오류 발생: {e}")


def run_09():
    try:
        # 저장할 Excel 경로
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path)

        # MAP_PAD 시트 존재 확인
        if "MAP_PAD" not in wb.sheetnames:
            print("❌ 'MAP_PAD' 시트를 찾을 수 없습니다.")
            return

        ws_map = wb["MAP_PAD"]

        # MAP_PAD 시트에서 A~E열 데이터 추출 (1행: 헤더)
        data = []
        for row in ws_map.iter_rows(min_row=1, max_col=5, values_only=True):
            data.append(row)

        # DataFrame 변환 및 헤더 분리
        df = pd.DataFrame(data[1:], columns=data[0])

        # 필터링
        df_fail = df[df["New_CODE"] == 4]              # E열이 4인 경우
        df_repair = df[df["New_CODE"].isin([0, 4])]    # E열이 0 또는 4인 경우

        # 스타일 지정
        align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # # Fail_Pin 시트 생성
        # if "Fail_Pin" in wb.sheetnames:
        #     del wb["Fail_Pin"]
        # ws_fail = wb.create_sheet("Fail_Pin")

        # for r_idx, row in enumerate(dataframe_to_rows(df_fail, index=False, header=True), start=1):
        #     for c_idx, value in enumerate(row, start=1):
        #         cell = ws_fail.cell(row=r_idx, column=c_idx, value=value)
        #         cell.alignment = align_center
        #         cell.border = thin_border

        # Repair_Pin 시트 생성
        if "Repair_Pin" in wb.sheetnames:
            del wb["Repair_Pin"]
        ws_repair = wb.create_sheet("Repair_Pin")

        for r_idx, row in enumerate(dataframe_to_rows(df_repair, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_repair.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = align_center
                cell.border = thin_border

        # 저장
        wb.save(output_excel_path)
        print("✅ run_09: 'Repair_Pin' 시트 생성 완료")

    except Exception as e:
        print(f"❌ run_09에서 오류 발생: {e}")


def run_10():
    try:
        # 저장할 Excel 경로
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = openpyxl.load_workbook(output_excel_path, data_only=True)

        # MAP_PAD 시트 존재 확인
        if "MAP_PAD" not in wb.sheetnames:
            print("❌ 'MAP_PAD' 시트를 찾을 수 없습니다.")
            return

        ws = wb["MAP_PAD"]

        # 헤더 작성
        ws["F1"] = "[MAP_PAD]"

        for row in range(2, ws.max_row + 1):
            dut = ws[f"A{row}"].value
            pad = ws[f"B{row}"].value
            New_code = ws[f"E{row}"].value

            if New_code == 0:
                ws[f"F{row}"] = f"{dut}_{pad}=0"
            elif New_code == 4:
                ws[f"F{row}"] = f"{dut}_{pad}=1"
            elif New_code == 10:
                ws[f"F{row}"] = f"{dut}_{pad}=3"
            else:   # 나머지 3
                ws[f"F{row}"] = f"{dut}_{pad}={New_code}"

        # 저장
        wb.save(output_excel_path)
        wb.close()
        print("✅ run_10: 'MAP_PAD' 시트의 F열 문자열 구성 완료")

    except Exception as e:
        print(f"❌ run_10에서 오류 발생: {e}")


def run_11():
    try:
        # 저장할 Excel 경로
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = openpyxl.load_workbook(output_excel_path, data_only=True)

        # 시트 확인
        if "MAP_DUT" not in wb.sheetnames or "MAP_PAD" not in wb.sheetnames:
            print("❌ 'MAP_DUT' 또는 'MAP_PAD' 시트를 찾을 수 없습니다.")
            return

        ws_dut = wb["MAP_DUT"]
        ws_pad = wb["MAP_PAD"]

        # MAP_PAD 시트: DUT -> E값 리스트 만들기
        pad_data = {}
        for row in range(2, ws_pad.max_row + 1):
            dut = ws_pad[f"A{row}"].value
            e_val = str(ws_pad[f"E{row}"].value)
            if dut not in pad_data:
                pad_data[dut] = []
            pad_data[dut].append(e_val)

        # MAP_DUT 시트: 헤더 추가
        ws_dut["C1"] = "New_Pad"

        # 각 DUT 검사 및 C열 기록
        for row in range(2, ws_dut.max_row + 1):
            dut = ws_dut[f"A{row}"].value
            codes = pad_data.get(dut, [])

            if "4" in codes:
                ws_dut[f"C{row}"] = 1
            elif (("3" in codes) or ("10" in codes)) and ("0" in codes):
                ws_dut[f"C{row}"] = 6
            elif "0" in codes:
                ws_dut[f"C{row}"] = 0
            else:  # 0 or 10 있는 경우
                ws_dut[f"C{row}"] = 3

        # 저장
        wb.save(output_excel_path)
        wb.close()
        print("✅ run_11: 'MAP_DUT' 시트에 C열(New_Pad) 값 작성 완료")

    except Exception as e:
        print(f"❌ run_11에서 오류 발생: {e}")


def run_12():
    try:
        # 저장할 Excel 경로
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = openpyxl.load_workbook(output_excel_path)
        sheet_dut = wb["MAP_DUT"]

        # 헤더 작성
        sheet_dut["D1"] = "[MAP_DUT]"

        # 각 행에 결과 문자열 작성
        for row in range(2, sheet_dut.max_row + 1):
            dut = sheet_dut[f"A{row}"].value
            new_pad = sheet_dut[f"C{row}"].value

            if dut is not None and new_pad not in [None, ""]:
                result = f"{dut}={new_pad}"
                sheet_dut[f"D{row}"] = result
            else:
                sheet_dut[f"D{row}"] = ""

        # 저장 및 닫기
        wb.save(output_excel_path)
        wb.close()
        print("✅ run_12: 'MAP_DUT' 시트의 D열 문자열 구성 완료")

    except Exception as e:
        print(f"❌ run_12에서 오류 발생: {e}")


def run_13():   ''' ### _new.csv _new.st2 삭제 ###
    import openpyxl
    import shutil
    import os
    import stat

    try:
        # 경로 설정
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        st2_output_path = os.path.splitext(st2_file_path)[0] + '_new.st2'
        csv_new_path = os.path.splitext(csv_file_path)[0] + '_new.csv'

        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(output_excel_path, data_only=True)
        sheet_pad = wb["MAP_PAD"]
        sheet_dut = wb["MAP_DUT"]

        # MAP_DUT 시트의 D열 수집
        data_1 = []
        for row in range(1, sheet_dut.max_row + 1):
            val = sheet_dut[f"D{row}"].value
            if val:
                data_1.append(str(val).strip())

        # MAP_PAD 시트의 F열 수집
        data_3 = []
        for row in range(1, sheet_pad.max_row + 1):
            val = sheet_pad[f"F{row}"].value
            if val:
                data_3.append(str(val).strip())

        # st2 리스트 구성
        data_2 = [None]  # 구분자
        st2 = data_1 + data_2 + data_3

        # .st2 파일 저장
        with open(st2_output_path, 'w', encoding='utf-8') as f:
            for val in st2:
                f.write("" if val is None else str(val))
                f.write("\n")

        print(f"✅ ST2 파일 저장 완료: {st2_output_path}")

        # 기존 CSV 복사 대상 파일이 있다면 삭제
        if os.path.exists(csv_new_path):
            try:
                os.chmod(csv_new_path, stat.S_IWRITE)  # 쓰기 가능하도록 권한 수정
                os.remove(csv_new_path)
                print(f"✅ 기존 CSV 파일 삭제: {csv_new_path}")
            except Exception as e:
                print(f"❌ 기존 CSV 파일 삭제 실패: {e}")
                return

        # CSV 복사 수행
        if os.path.exists(csv_file_path):
            shutil.copy(csv_file_path, csv_new_path)
            print(f"✅ CSV 파일 복사 완료: {csv_new_path}")
        else:
            print(f"❌ 원본 CSV 파일이 존재하지 않습니다: {csv_file_path}")

        # 엑셀 파일 닫기
        wb.close()

    except Exception as e:
        print(f"❌ run_13에서 오류 발생: {e}")
'''



##########################################################
def run_14():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'

        def read_sheet(filepath):
            ext = os.path.splitext(filepath)[1].lower()
            if ext == '.csv':
                try:
                    df = pd.read_csv(filepath, encoding='cp949', header=None)
                except UnicodeDecodeError:
                    df = pd.read_csv(filepath, encoding='utf-8', header=None)
            elif ext in ['.xlsx', '.xls']:
                df = pd.read_excel(filepath, header=None)
            else:
                raise ValueError(f"❌ 지원되지 않는 파일 형식: {filepath}")

            df = df.dropna(how='all')
            df = df.loc[~(df.map(lambda x: str(x).strip()).eq('').all(axis=1))]
            return df

        df_1200 = read_sheet(file_1200_path)
        df_1300 = read_sheet(file_1300_path)

        wb_out = openpyxl.load_workbook(output_excel_path)
        sheet_name = 'Repair_sheet'
        if sheet_name in wb_out.sheetnames:
            ws_out = wb_out[sheet_name]
        else:
            ws_out = wb_out.create_sheet(sheet_name)

        for row in ws_out.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        for i, row in enumerate(df_1300.itertuples(index=False, name=None), start=3):
            for j, val in enumerate(row, start=1):
                ws_out.cell(row=i, column=j, value=val)

        for i, row in enumerate(df_1200.itertuples(index=False, name=None), start=3):
            for j, val in enumerate(row, start=8):
                ws_out.cell(row=i, column=j, value=val)

        ws_out.merge_cells("A1:G1")
        ws_out.merge_cells("H1:N1")
        ws_out["A1"] = "1300 Repair_Sheet"
        ws_out["H1"] = "1200 Repair_Sheet"

        headers_1300 = ["No.", "DUT", "Type", "PAD", "S/L", "X 좌표", "Y 좌표"]
        headers_1200 = ["No.", "DUT", "정/역", "PAD", "S/L", "X 좌표", "Y 좌표"]
        for col, value in enumerate(headers_1300, start=1):
            ws_out.cell(row=2, column=col, value=value)
        for col, value in enumerate(headers_1200, start=8):
            ws_out.cell(row=2, column=col, value=value)

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for col in range(1, 8):
            cell = ws_out.cell(row=1, column=col)
            cell.alignment = center_align
            cell.border = thin_border
        for col in range(8, 15):
            cell = ws_out.cell(row=1, column=col)
            cell.alignment = center_align
            cell.border = thin_border

        wb_out.save(output_excel_path)
        print(f"'Repair_sheet' 시트 생성 및 데이터 입력 완료.")

    except Exception as e:
        print(f"run_09 오류 발생: {e}")



def run_15():
    try:
        # === 경로 설정 ===
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        source_sheet = 'Repair_sheet'
        filter_sheet = 'Repair_sheet_filter'

        # === 데이터 읽기 ===
        df = pd.read_excel(output_excel_path, sheet_name=source_sheet, header=1, usecols='A:N')

        # === 필터링 ===
        if len(sl_values) == 1 and sl_values[0].upper() == "S/L":
            filtered_df = df[df['S/L'].notna() & (df['S/L'].astype(str).str.strip() != '')]
        else:
            filtered_df = df[df['S/L'].isin(sl_values)]

        # === 엑셀 파일 열기 ===
        wb = load_workbook(output_excel_path)

        # === 기존 시트 삭제 ===
        if filter_sheet in wb.sheetnames:
            del wb[filter_sheet]

        # === 새 시트 생성 ===
        ws = wb.create_sheet(filter_sheet)

        # === 타이틀 작성 ===
        ws.merge_cells("A1:G1")
        ws.merge_cells("H1:N1")
        ws["A1"] = "1300 Repair_Sheet"
        ws["H1"] = "1200 Repair_Sheet"

        # === 셀 스타일 정의 ===
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col in range(1, 8):
            cell = ws.cell(row=1, column=col)
            cell.alignment = center_align
            cell.border = thin_border
        for col in range(8, 15):
            cell = ws.cell(row=1, column=col)
            cell.alignment = center_align
            cell.border = thin_border

        # === 필터링된 데이터 작성 ===
        for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # === 저장 ===
        wb.save(output_excel_path)
        print(f"'Repair_sheet_filter' 시트에 '{', '.join(sl_values)}' 필터 결과 저장 완료.")

    except Exception as e:
        print(f"❌ run_15 오류 발생: {e}")


def run_16():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path)
        sheet_name = '1300-1200_Repair'

        # 시트 생성 또는 선택
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # 1행 병합 및 타이틀 입력
        ws.merge_cells("A1:C1")
        ws.merge_cells("D1:F1")
        ws["A1"] = "1300 Repair_Sheet"
        ws["D1"] = "1200 Repair_Sheet"

        # 2행 헤더 입력
        headers = ["DUT", "PAD", "fail", "정/역", "DUT", "PAD"]
        for col, value in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=value)

        # ✅ Repair_Pin 시트에서 A, B, E열 읽기
        df_pin = pd.read_excel(output_excel_path, sheet_name='Repair_Pin', usecols=[0, 1, 4], header=0)
        df_pin = df_pin.dropna(how='all')
        df_pin = df_pin.iloc[0:]  # Excel 기준 3행부터 가져옴 (0-indexed → 2)

        # ✅ 3행부터 A, B, C열에 작성
        for row_idx, (dut, pad, fail) in enumerate(df_pin.itertuples(index=False, name=None), start=3):
            ws.cell(row=row_idx, column=1, value=dut)   # A열
            ws.cell(row=row_idx, column=2, value=pad)   # B열
            ws.cell(row=row_idx, column=3, value=fail)  # C열

        wb.save(output_excel_path)
        print(f"'1300-1200_Repair' 시트에 Repair_Pin 데이터가 추가되었습니다.")

    except Exception as e:
        print(f"run_16 오류 발생: {e}")



def run_17():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path)
        repair_sheet = '1300-1200_Repair'
        filter_sheet = 'Repair_sheet_filter'

        if repair_sheet not in wb.sheetnames or filter_sheet not in wb.sheetnames:
            print(f"❌ 필요한 시트가 존재하지 않습니다.")
            return

        ws = wb[repair_sheet]
        ws_filter = wb[filter_sheet]

        # 가운데 정렬 + 테두리 스타일 정의
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 1행 병합 영역 스타일
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.alignment = center_align
            cell.border = thin_border
        for col in range(4, 7):
            cell = ws.cell(row=1, column=col)
            cell.alignment = center_align
            cell.border = thin_border

        # 2행 헤더 스타일
        for col in range(1, 7):
            cell = ws.cell(row=2, column=col)
            cell.alignment = center_align
            cell.border = thin_border

        # ✅ Repair_sheet_filter 데이터 읽기 (필요한 열만)
        filter_data = {}
        for row in ws_filter.iter_rows(min_row=2, values_only=True):  # 헤더 제외
            key = (row[1], row[3])  # B열, D열
            value = (row[9], row[8], row[10])  # J열, I열, K열
            filter_data[key] = value

        # ✅ 1300-1200_Repair 데이터와 매칭하여 D/E/F열에 입력
        row_idx = 3
        while True:
            dut = ws.cell(row=row_idx, column=1).value  # A열
            pad = ws.cell(row=row_idx, column=2).value  # B열
            if dut is None and pad is None:
                break  # 끝

            match = filter_data.get((dut, pad))
            if match:
                ws.cell(row=row_idx, column=4, value=match[0])  # D열 ← J
                ws.cell(row=row_idx, column=5, value=match[1])  # E열 ← I
                ws.cell(row=row_idx, column=6, value=match[2])  # F열 ← K

            row_idx += 1

        wb.save(output_excel_path)
        print(f"'1300-1200_Repair' 시트에 데이터가 적용되었습니다.")

    except Exception as e:
        print(f"run_17 오류 발생: {e}")


def run_18():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path)

        src_sheet = '1300-1200_Repair'
        print_sheet = 'PRINT'

        if src_sheet not in wb.sheetnames:
            print(f"❌ '{src_sheet}' 시트가 존재하지 않습니다.")
            return

        ws_src = wb[src_sheet]

        if print_sheet in wb.sheetnames:
            del wb[print_sheet]
        ws_print = wb.create_sheet(print_sheet)
        ws_print.sheet_properties.tabColor = "FFFF00"

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # B열이 X/Y/Z 3열로 분리되어 전체 열이 2칸 오른쪽으로 이동
        ws_print.merge_cells("A1:E2")
        ws_print.merge_cells("A3:E3")
        ws_print.merge_cells("F3:H3")
        ws_print.merge_cells("G1:H1")
        ws_print.merge_cells("G2:H2")
        ws_print.merge_cells("I3:J3")
        ws_print.merge_cells("J1:L1")
        ws_print.merge_cells("J2:L2")
        ws_print.merge_cells("K3:O3")
        ws_print.merge_cells("M1:N1")
        ws_print.merge_cells("M2:N2")

        ws_print["A1"] = ", ".join(sl_values) if sl_values else ""
        ws_print["F1"] = "작성자"
        ws_print["F2"] = "Probe"
        ws_print["F3"] = "1300 Repair_Sheet"
        ws_print["I1"] = "제품 No"
        ws_print["I2"] = "설비 No"
        ws_print["I3"] = "1200 Repair_Sheet"
        ws_print["K3"] = "Inspection"
        ws_print["M1"] = "기판"
        ws_print["M2"] = "Fail X Y Z"
        ws_print["O1"] = " / "
        ws_print["O2"] = '=SUM(B5:D1000)&" / "&SUM(B5:B1000)&" / "&SUM(C5:C1000)&" / "&SUM(D5:D1000)'
        ws_print["Q1"] = "Fail"
        ws_print["Q2"] = "X ="
        ws_print["Q3"] = "Y ="
        ws_print["Q4"] = "Z ="
        ws_print["R1"] = '=SUM(B5:D1000)'
        ws_print["R2"] = '=SUM(B5:B1000)'
        ws_print["R3"] = '=SUM(C5:C1000)'
        ws_print["R4"] = '=SUM(D5:D1000)'
        ws_print["T1"] = '전체 Pin'
        ws_print["T2"] = 'Fail'
        ws_print["T3"] = 'Skip'
        ws_print["T4"] = 'Pass'
        ws_print["A3"] = datetime.today().strftime("%y-%m-%d")

        bold_font = Font(bold=True)
        ws_print["F1"].font = bold_font
        ws_print["F2"].font = bold_font
        ws_print["F3"].font = bold_font
        ws_print["I1"].font = bold_font
        ws_print["I2"].font = bold_font
        ws_print["I3"].font = bold_font
        ws_print["M1"].font = bold_font
        ws_print["M2"].font = bold_font
        ws_print["K3"].font = bold_font

        # X/Y/Z 3열 분리로 헤더도 2칸 이동
        headers = ["No", "X", "Y", "Z", "정/역", "DUT", "PAD", "빈 PAD", "DUT", "PAD", "냉납", "Burnt", "눌림", "패드X", "내용"]
        for col, value in enumerate(headers, start=1):
            cell = ws_print.cell(row=4, column=col, value=value)
            cell.alignment = center_align
            cell.border = thin_border

        for row in range(1, 4):
            for col in range(1, 16):  # A ~ O
                cell = ws_print.cell(row=row, column=col)
                cell.alignment = center_align
                cell.border = thin_border

        print_row = 5
        pad_B_list, pad_C_list = [], []
        pad_F_list, pad_G_list = [], []

        first_dut = None
        first_rev = None
        first_dut2 = None

        thin_black = Side(style="thin", color="000000")

        for row in range(1, 5):
            for cols in (("Q", "R"), ("T", "U")):
                for col in cols:
                    cell = ws_print[f"{col}{row}"]
                    cell.alignment = center_align
                    cell.border = Border(
                        top=thin_black,
                        bottom=thin_black,
                        left=thin_black if col == cols[0] else None,
                        right=thin_black if col == cols[-1] else None
                    )

        def format_pad_list(pads):
            if not pads:
                return ""
            pads = sorted(set(pads))
            ranges = []
            start = pads[0]
            end = pads[0]
            for p in pads[1:]:
                if p == end + 1:
                    end = p
                else:
                    if start == end:
                        ranges.append(f"{start}")
                    else:
                        ranges.append(f"{start} ~ {end}")
                    start = end = p
            if start == end:
                ranges.append(f"{start}")
            else:
                ranges.append(f"{start} ~ {end}")
            return ", ".join(ranges)

        for row_idx in range(3, ws_src.max_row + 1):
            dut = ws_src.cell(row=row_idx, column=1).value
            pad = ws_src.cell(row=row_idx, column=2).value
            fail = ws_src.cell(row=row_idx, column=3).value
            rev = ws_src.cell(row=row_idx, column=4).value
            dut2 = ws_src.cell(row=row_idx, column=5).value
            pad2 = ws_src.cell(row=row_idx, column=6).value

            next_dut = ws_src.cell(row=row_idx + 1, column=1).value if row_idx + 1 <= ws_src.max_row else None
            next_pad = ws_src.cell(row=row_idx + 1, column=2).value if row_idx + 1 <= ws_src.max_row else None
            next_fail = ws_src.cell(row=row_idx + 1, column=3).value if row_idx + 1 <= ws_src.max_row else None
            next_rev = ws_src.cell(row=row_idx + 1, column=4).value if row_idx + 1 <= ws_src.max_row else None
            next_dut2 = ws_src.cell(row=row_idx + 1, column=5).value if row_idx + 1 <= ws_src.max_row else None

            if first_dut is None:
                first_dut = dut
                first_rev = rev
                first_dut2 = dut2

            if fail == 0:
                pad_C_list.append(pad)
                pad_G_list.append(pad2)
            elif fail == 4:
                pad_B_list.append(pad)
                pad_F_list.append(pad2)

            is_group_end = (
                (next_dut != dut or next_rev != rev or next_dut2 != dut2) or
                (next_fail not in [0, 4]) or
                (next_dut is None)
            )

            if is_group_end:
                # 열 번호 +2 (X/Y/Z 분리로 인한 이동)
                ws_print.cell(row=print_row, column=5, value=first_rev).alignment = center_align   # 정/역
                ws_print.cell(row=print_row, column=6, value=first_dut).alignment = center_align   # DUT
                ws_print.cell(row=print_row, column=9, value=first_dut2).alignment = center_align  # DUT2
                ws_print.cell(row=print_row, column=7, value=format_pad_list(pad_B_list)).alignment = center_align   # PAD (Fail=4)
                ws_print.cell(row=print_row, column=8, value=format_pad_list(pad_C_list)).alignment = center_align   # 빈 PAD (Fail=0)
                ws_print.cell(row=print_row, column=10, value=format_pad_list(pad_F_list)).alignment = center_align  # PAD2 (Fail=4)

                print_row += 1
                pad_B_list.clear()
                pad_C_list.clear()
                pad_F_list.clear()
                pad_G_list.clear()
                first_dut = None
                first_rev = None
                first_dut2 = None

        gray_thin_side = Side(style="thin", color="BFBFBF")
        black_thin_side = Side(style="thin", color="000000")

        for row in range(5, print_row):
            for col in range(1, 16):  # A ~ O
                cell = ws_print.cell(row=row, column=col)
                bottom = gray_thin_side
                if col in [9, 11]:  # I열(DUT2 시작), K열(냉납 시작)
                    left = black_thin_side
                elif col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
                    left = gray_thin_side
                else:
                    left = Side(border_style=None)
                cell.border = Border(left=left, bottom=bottom)

        fit_center_align = Alignment(
            horizontal="center",
            vertical="center",
            shrink_to_fit=True
        )
        center_only_align = Alignment(horizontal="center", vertical="center")

        center_cols = [2, 3, 4, 12, 13, 14, 15]  # B(X), C(Y), D(Z), L(냉납), M(Burnt), N(눌림), O(패드X)
        for row in range(5, print_row):
            for col in center_cols:
                cell = ws_print.cell(row=row, column=col)
                if cell.value is not None and cell.value != "":
                    cell.alignment = center_only_align

        target_cols = [7, 8, 10, 11]  # G(PAD), H(빈PAD), J(PAD2), K(냉납)
        for row in range(5, print_row):
            for col in target_cols:
                cell = ws_print.cell(row=row, column=col)
                if cell.value is not None and cell.value != "":
                    cell.alignment = fit_center_align

        # MAP_PAD E열 카운트
        e_col_4 = 0  # Fail (E=4)
        e_col_0 = 0  # Skip (E=0)
        if "MAP_PAD" in wb.sheetnames:
            ws_map = wb["MAP_PAD"]
            for r in range(2, ws_map.max_row + 1):
                val = ws_map.cell(r, 5).value
                if val == 4:
                    e_col_4 += 1
                elif val == 0:
                    e_col_0 += 1

        # T열 옆 U열에 통계 수치 기록
        ws_print["U1"] = len([v for v in list_block if isinstance(v, (int, float))])  # 전체 Pin
        ws_print["U2"] = e_col_4            # Fail = MAP_PAD E=4 개수
        ws_print["U3"] = e_col_0            # Skip = MAP_PAD E=0 개수
        ws_print["U4"] = '=U1-U2-U3'        # Pass = 전체 - Fail - Skip

        print_end_row = print_row - 1
        ws_print.print_area = f"A1:O{print_end_row}"
        ws_print.print_title_rows = '1:4'
        ws_print.page_setup.paperSize = ws_print.PAPERSIZE_A4
        ws_print.page_setup.fitToPage = True
        ws_print.page_setup.fitToWidth = 1
        ws_print.page_setup.fitToHeight = 0
        ws_print.page_setup.orientation = ws_print.ORIENTATION_LANDSCAPE
        ws_print.oddFooter.center.text = "페이지 &P"

        ws_print.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.4,
            bottom=0.6,
            header=0.3,
            footer=0.2
        )

        ws_print.column_dimensions["A"].width = 6.0
        ws_print.column_dimensions["B"].width = 4.0   # X
        ws_print.column_dimensions["C"].width = 4.0   # Y
        ws_print.column_dimensions["D"].width = 4.0   # Z
        ws_print.column_dimensions["E"].width = 6.0   # 정/역
        ws_print.column_dimensions["G"].width = 17.0  # PAD
        ws_print.column_dimensions["H"].width = 17.0  # 빈 PAD
        ws_print.column_dimensions["J"].width = 17.0  # PAD2
        ws_print.column_dimensions["K"].width = 6.0   # 냉납
        ws_print.column_dimensions["L"].width = 6.0   # Burnt
        ws_print.column_dimensions["M"].width = 6.0   # 눌림
        ws_print.column_dimensions["N"].width = 6.0   # 패드X
        ws_print.column_dimensions["O"].width = 34.0  # 내용
        ws_print.column_dimensions["U"].width = 8.0   # 통계 수치

        wb.save(output_excel_path)
        print(f"✅ 'PRINT' 시트가 생성되었고, 테두리 및 열 너비가 설정되었습니다.")

    except Exception as e:
        print(f"❌ run_18 오류 발생: {e}")



def main():
    try:
        print("01 실행 중")
        run_01()
        print("02 실행 중")
        run_02()
        print("03 실행 중")
        run_03()
        print("04 실행 중")
        run_04()
        print("05 실행 중")
        run_05()
        print("06 실행 중")
        run_06()
        print("07 실행 중")
        run_07()
        print("08 실행 중")
        run_08()
        print("09 실행 중")
        run_09()
        print("10 실행 중")
        run_10()
        print("11 실행 중")
        run_11()
        print("12 실행 중")
        run_12()
        print("13 실행 중")
        run_13()
        print("14 실행 중")
        run_14()
        print("15 실행 중")
        run_15()
        print("16 실행 중")
        run_16()
        print("17 실행 중")
        run_17()
        print("18 실행 중")
        run_18()
        print("모든 스크립트 실행 완료!")

    except Exception as e:
        with open("error_log.txt", "w", encoding="utf-8") as f:
            traceback.print_exc(file=f)
        print("에러가 발생했습니다. error_log.txt를 확인하세요.")

if __name__ == "__main__":
    main()