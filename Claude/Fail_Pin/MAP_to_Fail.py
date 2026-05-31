# ============================================================
# MAP_to_Fail.py  —  Fail PIN+MAP  +  XYZ Fail  통합 도구
# ============================================================

import os
import sys
import shutil
import stat
import traceback
import calendar
import threading
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
from pathlib import Path
import ctypes
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ============================================================
# FAIL_PIN+MAP — 전역 상태
# ============================================================
st2_file_path  = ""
csv_file_path  = ""
file_1200_path = ""
file_1300_path = ""
sl_values      = []
type_list      = []
type_lists     = {}
list_block     = []
stop_event     = threading.Event()


class _StopExecution(BaseException):
    pass


# ============================================================
# Console 리다이렉터
# ============================================================
class _ConsoleRedirect:
    def __init__(self, text_widget):
        self._w = text_widget

    def write(self, text):
        def _do():
            try:
                self._w.configure(state="normal")
                self._w.insert("end", text)
                total = int(self._w.index("end-1c").split(".")[0])
                if total > 500:
                    self._w.delete("1.0", f"{total - 500}.0")
                self._w.see("end")
                self._w.configure(state="disabled")
            except Exception:
                pass
        try:
            self._w.after(0, _do)
        except Exception:
            pass

    def flush(self):
        pass


# ============================================================
# FAIL_PIN+MAP — 처리 함수들
# ============================================================

def parse_sl_input(sl_input):
    values = []
    for v in sl_input.split(','):
        val = v.strip().lower()
        if val == 'l':
            values.append('Long')
        elif val == 's':
            values.append('Short')
        elif val:
            values.append(v.strip())
    return values


def run_01():
    if not st2_file_path.endswith('.st2'):
        print("'.st2' 확장자의 파일만 지원합니다.")
        return
    elif not os.path.exists(st2_file_path):
        print(f"파일이 존재하지 않습니다: {st2_file_path}")
        return
    try:
        with open(st2_file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
        cleaned_lines = [line.strip() for line in lines]
        df_st2 = pd.DataFrame(cleaned_lines)

        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            df_st2.to_excel(writer, index=False, header=False, sheet_name='st2')
        print(f"Excel 파일로 저장 완료: {output_excel_path}")

        if not csv_file_path.endswith('.csv'):
            print("'.csv' 확장자의 파일만 지원합니다.")
            return
        elif not os.path.exists(csv_file_path):
            print(f"CSV 파일이 존재하지 않습니다: {csv_file_path}")
            return

        df_csv = pd.read_csv(csv_file_path, header=None)
        with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_csv.to_excel(writer, index=False, header=False, sheet_name='CSV')
        print("CSV 데이터를 'CSV' 시트로 추가 완료.")
    except Exception as e:
        print(f"오류 발생: {e}")


def run_02():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path)
        if 'CSV' not in wb.sheetnames:
            print("시트 'CSV'이(가) 존재하지 않습니다.")
            return
        ws_source = wb['CSV']

        target_row = None
        for row in ws_source.iter_rows(min_col=1, max_col=1):
            val = row[0].value
            if isinstance(val, str) and "pin coodination" in val.lower():
                target_row = row[0].row
                break
        if target_row is None:
            print("'pin coodination'이 포함된 셀을 찾을 수 없습니다.")
            return

        new_sheet_name = 'pin coodination'
        if new_sheet_name not in wb.sheetnames:
            ws_target = wb.create_sheet(new_sheet_name)
        else:
            ws_target = wb[new_sheet_name]

        for r_idx, row in enumerate(ws_source.iter_rows(min_row=target_row), 1):
            for c_idx, cell in enumerate(row, 1):
                ws_target.cell(row=r_idx, column=c_idx, value=cell.value)

        bold_font = Font(bold=True)
        for col, header in enumerate(["Type", "PAD 거리", "정역구분", "거리함수", "끝 번호", "블럭"], start=7):
            cell = ws_target.cell(row=1, column=col, value=header)
            cell.font = bold_font

        wb.save(output_excel_path)
        print(f"'{new_sheet_name}' 시트에 헤더 적용 완료.")
    except Exception as e:
        print(f"오류 발생: {e}")


def run_03():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        sheet_to_load = 'pin coodination'
        df = pd.read_excel(output_excel_path, sheet_name=sheet_to_load, header=None, skiprows=1)

        df['H_PAD_거리'] = (df[3] - df[3].shift(-1)).fillna(0).astype(int)
        df['I_정역구분'] = (df[4] - df[4].shift(-1)).fillna(0).astype(int)

        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb[sheet_to_load]
        for df_idx, row_data in df.iterrows():
            excel_row = df_idx + 2
            ws.cell(row=excel_row, column=8, value=row_data['H_PAD_거리'])
            ws.cell(row=excel_row, column=9, value=row_data['I_정역구분'])
        wb.save(output_excel_path)
        print(f"'{sheet_to_load}' 시트의 H열과 I열 계산 및 업데이트 완료.")
    except Exception as e:
        print(f"run_03 오류 발생: {e}")


def run_04():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        sheet_to_load = 'pin coodination'
        df = pd.read_excel(output_excel_path, sheet_name=sheet_to_load, header=None, skiprows=1)

        df['J_거리함수'] = (df[7].abs() > 800) | (df[8].abs() > 1000)
        df['K_끝번호']   = df.apply(lambda r: r[0] if r['J_거리함수'] else "-", axis=1)
        df['L_블럭']     = 1

        for i in range(1, len(df)):
            if df.loc[i, 6] == df.loc[i-1, 6]:
                df.loc[i, 'L_블럭'] = df.loc[i-1, 'L_블럭'] + 1 if df.loc[i-1, 'K_끝번호'] == "-" else 1
            else:
                df.loc[i, 'L_블럭'] = 1

        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb[sheet_to_load]
        for df_idx, row_data in df.iterrows():
            excel_row = df_idx + 2
            ws.cell(row=excel_row, column=10, value=row_data['J_거리함수'])
            ws.cell(row=excel_row, column=11, value=row_data['K_끝번호'])
            ws.cell(row=excel_row, column=12, value=row_data['L_블럭'])
        wb.save(output_excel_path)
        print(f"'{sheet_to_load}' 시트의 J, K, L열 계산 및 업데이트 완료.")
    except Exception as e:
        print(f"run_04 오류 발생: {e}")


def run_05():
    try:
        global type_list, type_lists
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path, data_only=True)

        if "pin coodination" not in wb.sheetnames:
            print("'pin coodination' 시트를 찾을 수 없습니다.")
            return
        ws = wb["pin coodination"]

        target_row = None
        for row in ws.iter_rows(min_col=1, max_col=1):
            if isinstance(row[0].value, str) and "pin coodination" in row[0].value.lower():
                target_row = row[0].row
                break
        if target_row is None:
            print("'pin coodination' 텍스트가 포함된 셀을 찾을 수 없습니다.")
            return

        data_start_row = target_row + 1
        max_row = ws.max_row

        type_values = set()
        for r in range(data_start_row, max_row + 1):
            g_val = ws.cell(row=r, column=7).value
            if g_val:
                type_values.add(str(g_val).strip())

        type_list  = sorted(type_values)
        type_lists = {f"type_list_{i}": [] for i in range(len(type_list))}

        for r in range(data_start_row, max_row + 1):
            g_val = ws.cell(row=r, column=7).value
            l_val = ws.cell(row=r, column=12).value
            if g_val is None or l_val is None:
                continue
            for i, type_name in enumerate(type_list):
                if str(g_val).strip() == type_name:
                    type_lists[f"type_list_{i}"].append(l_val)

        if "Type" in wb.sheetnames:
            del wb["Type"]
        ws_type = wb.create_sheet("Type")
        for i, type_name in enumerate(type_list):
            ws_type.cell(row=1, column=i + 1, value=type_name)
        for i, key in enumerate(type_lists.keys()):
            for j, val in enumerate(type_lists[key], start=2):
                ws_type.cell(row=j, column=i + 1, value=val)

        wb.save(output_excel_path)
        print("'Type' 시트 생성 및 L열 값 분류 완료.")
        print("Type List:", type_list)
    except Exception as e:
        print(f"오류 발생: {e}")


def run_06():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path, data_only=True)
        ws = wb.active
        ws.title = "st2"

        map_dut_sheet = wb.create_sheet(title="MAP_DUT")
        map_pad_sheet = wb.create_sheet(title="MAP_PAD")

        dut_list = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if cell.value:
                    if cell.value == '[MAP_PAD]':
                        break
                    split_values = cell.value.split('=')
                    dut_list.append(split_values if len(split_values) == 2 else [split_values[0], ''])
            if cell.value == '[MAP_PAD]':
                break

        map_dut_sheet.append(['DUT', 'PAD'])
        for i, (d1, d2) in enumerate(dut_list, start=2):
            map_dut_sheet.cell(row=i, column=1, value=d1)
            map_dut_sheet.cell(row=i, column=2, value=d2)

        map_pad_processing = False
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if cell.value:
                    if cell.value == '[MAP_PAD]':
                        map_pad_processing = True
                        continue
                    if map_pad_processing:
                        row_values = []
                        for value in cell.value.split('_'):
                            row_values.extend(value.split('='))
                        if row_values[0] != 'DUT':
                            map_pad_sheet.append(row_values)

        map_pad_sheet.insert_rows(1)
        map_pad_sheet.cell(row=1, column=1, value='DUT')
        map_pad_sheet.cell(row=1, column=2, value='PAD')
        map_pad_sheet.cell(row=1, column=3, value='CODE')
        wb.save(output_excel_path)
    except Exception as e:
        print(f"오류 발생: {e}")


def run_07():
    try:
        global list_block
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path, data_only=True)
        csv_ws  = wb["CSV"]
        type_ws = wb["Type"]

        target_row = None
        for row in csv_ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == "substrate dut coodination":
                    target_row = cell.row
                    break
            if target_row:
                break
        if not target_row:
            print("CSV 시트에서 'substrate dut coodination'을 찾을 수 없습니다.")
            return

        d_values = []
        r = target_row + 1
        while True:
            val = csv_ws.cell(row=r, column=4).value
            if val is None:
                break
            d_values.append(val)
            r += 1

        matched_cols = []
        for d_val in d_values:
            for col in range(1, type_ws.max_column + 1):
                header = type_ws.cell(row=1, column=col).value
                if header and str(header).strip() == str(d_val).strip():
                    matched_cols.append(col)
                    break

        if not matched_cols:
            print(f"Type 시트에서 {d_values} 중 매칭되는 열을 찾을 수 없습니다.")
            try:
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                messagebox.showerror("TYPE 오류", "CSV 파일에서 TYPE 매칭되게 수정해주세요")
                root.destroy()
            except Exception:
                pass
            sys.exit(1)

        list_block = []
        for col in matched_cols:
            for r in range(2, type_ws.max_row + 1):
                val = type_ws.cell(row=r, column=col).value
                if val is not None:
                    list_block.append(val)

        print(f"list_block 개수: {len([v for v in list_block if isinstance(v, (int, float))])}")

        if "MAP_PAD" not in wb.sheetnames:
            print("MAP_PAD 시트를 찾을 수 없습니다.")
            return

        map_pad_ws = wb["MAP_PAD"]
        map_pad_ws.cell(row=1, column=4, value="블럭")
        for i, val in enumerate(list_block):
            map_pad_ws.cell(row=2 + i, column=4, value=val)
        wb.save(output_excel_path)
        print("MAP_PAD 시트에 '블럭' 열과 list_block 데이터를 작성했습니다.")
    except Exception as e:
        print(f"run_07에서 오류 발생: {e}")


def run_08():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path)
        if "MAP_PAD" not in wb.sheetnames:
            print("MAP_PAD 시트를 찾을 수 없습니다.")
            return
        ws = wb["MAP_PAD"]

        data = [row for row in ws.iter_rows(min_row=2, min_col=3, max_col=4, values_only=True)]
        df = pd.DataFrame(data, columns=["CODE", "BLOCK"])
        df["CODE"]  = pd.to_numeric(df["CODE"],  errors="coerce").fillna(-1).astype(int)
        df["BLOCK"] = pd.to_numeric(df["BLOCK"], errors="coerce").fillna(-1).astype(int)
        df["New_CODE"]   = None
        df["BLOCK_DIFF"] = df["BLOCK"].diff().fillna(1)
        df["BLOCK_GROUP"] = (df["BLOCK_DIFF"] != 1).cumsum()

        for _, group in df.groupby("BLOCK_GROUP"):
            codes   = group["CODE"].tolist()
            indexes = group.index.tolist()
            for i, idx in enumerate(indexes):
                code = codes[i]
                if code in [3, 5, 6]:
                    df.at[idx, "New_CODE"] = 4 if 0 in codes[:i] else 3
                elif code in [4, 8, 9]:
                    is_last = (i == len(codes) - 1)
                    if i > 0 and codes[i-1] in [3, 5, 6] and is_last:
                        df.at[idx, "New_CODE"] = 10
                        continue
                    if 0 in codes[:i]:
                        df.at[idx, "New_CODE"] = 4
                        continue
                    result = 4
                    for nc in codes[i+1:]:
                        if nc in [3, 5, 6]:
                            result = 3
                            break
                        elif nc == 0:
                            break
                    df.at[idx, "New_CODE"] = result
                elif code == 0:
                    found = False
                    for j in range(i + 1, len(codes)):
                        if codes[j] != 0:
                            found = True
                            for k in range(i, j + 1):
                                df.at[indexes[k], "New_CODE"] = 4
                            break
                    if not found:
                        df.at[idx, "New_CODE"] = 0

        ws.cell(row=1, column=5, value="New_CODE")
        for i, val in enumerate(df["New_CODE"], start=2):
            ws.cell(row=i, column=5, value=int(val) if pd.notna(val) else "")
        wb.save(output_excel_path)
        print("✅ run_08: 'MAP_PAD' 시트의 E열 'New_CODE' 계산 완료")
    except Exception as e:
        print(f"❌ run_08에서 오류 발생: {e}")


def run_09():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = load_workbook(output_excel_path)
        if "MAP_PAD" not in wb.sheetnames:
            print("❌ 'MAP_PAD' 시트를 찾을 수 없습니다.")
            return
        ws_map = wb["MAP_PAD"]

        data = [row for row in ws_map.iter_rows(min_row=1, max_col=5, values_only=True)]
        df = pd.DataFrame(data[1:], columns=data[0])
        df_repair = df[df["New_CODE"].isin([0, 4])]

        align_center = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"),  bottom=Side(style="thin"))

        if "Repair_Pin" in wb.sheetnames:
            del wb["Repair_Pin"]
        ws_repair = wb.create_sheet("Repair_Pin")
        for r_idx, row in enumerate(dataframe_to_rows(df_repair, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_repair.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = align_center
                cell.border    = thin_border

        wb.save(output_excel_path)
        print("✅ run_09: 'Repair_Pin' 시트 생성 완료")
    except Exception as e:
        print(f"❌ run_09에서 오류 발생: {e}")


def run_10():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = openpyxl.load_workbook(output_excel_path, data_only=True)
        if "MAP_PAD" not in wb.sheetnames:
            print("❌ 'MAP_PAD' 시트를 찾을 수 없습니다.")
            return
        ws = wb["MAP_PAD"]
        ws["F1"] = "[MAP_PAD]"
        for row in range(2, ws.max_row + 1):
            dut      = ws[f"A{row}"].value
            pad      = ws[f"B{row}"].value
            new_code = ws[f"E{row}"].value
            if new_code == 0:
                ws[f"F{row}"] = f"{dut}_{pad}=0"
            elif new_code == 4:
                ws[f"F{row}"] = f"{dut}_{pad}=1"
            elif new_code == 10:
                ws[f"F{row}"] = f"{dut}_{pad}=3"
            else:
                ws[f"F{row}"] = f"{dut}_{pad}={new_code}"
        wb.save(output_excel_path)
        wb.close()
        print("✅ run_10: 'MAP_PAD' 시트의 F열 문자열 구성 완료")
    except Exception as e:
        print(f"❌ run_10에서 오류 발생: {e}")


def run_11():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = openpyxl.load_workbook(output_excel_path, data_only=True)
        if "MAP_DUT" not in wb.sheetnames or "MAP_PAD" not in wb.sheetnames:
            print("❌ 'MAP_DUT' 또는 'MAP_PAD' 시트를 찾을 수 없습니다.")
            return
        ws_dut = wb["MAP_DUT"]
        ws_pad = wb["MAP_PAD"]

        pad_data = {}
        for row in range(2, ws_pad.max_row + 1):
            dut   = ws_pad[f"A{row}"].value
            e_val = str(ws_pad[f"E{row}"].value)
            pad_data.setdefault(dut, []).append(e_val)

        ws_dut["C1"] = "New_Pad"
        for row in range(2, ws_dut.max_row + 1):
            dut   = ws_dut[f"A{row}"].value
            codes = pad_data.get(dut, [])
            if "4" in codes:
                ws_dut[f"C{row}"] = 1
            elif (("3" in codes) or ("10" in codes)) and ("0" in codes):
                ws_dut[f"C{row}"] = 6
            elif "0" in codes:
                ws_dut[f"C{row}"] = 0
            else:
                ws_dut[f"C{row}"] = 3

        wb.save(output_excel_path)
        wb.close()
        print("✅ run_11: 'MAP_DUT' 시트에 C열(New_Pad) 값 작성 완료")
    except Exception as e:
        print(f"❌ run_11에서 오류 발생: {e}")


def run_12():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb["MAP_DUT"]
        ws["D1"] = "[MAP_DUT]"
        for row in range(2, ws.max_row + 1):
            dut     = ws[f"A{row}"].value
            new_pad = ws[f"C{row}"].value
            ws[f"D{row}"] = f"{dut}={new_pad}" if (dut is not None and new_pad not in [None, ""]) else ""
        wb.save(output_excel_path)
        wb.close()
        print("✅ run_12: 'MAP_DUT' 시트의 D열 문자열 구성 완료")
    except Exception as e:
        print(f"❌ run_12에서 오류 발생: {e}")


def run_13():
    pass  # 비활성화 (_new.csv / _new.st2 삭제 기능)


def run_14():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'

        def read_sheet(filepath):
            ext = os.path.splitext(filepath)[1].lower()
            if ext == '.csv':
                try:
                    df = pd.read_csv(filepath, encoding='cp949', header=None)
                except UnicodeDecodeError:
                    df = pd.read_csv(filepath, encoding='utf-8', header=None)
            elif ext in ['.xlsx', '.xls']:
                df = pd.read_excel(filepath, header=None)
            else:
                raise ValueError(f"❌ 지원되지 않는 파일 형식: {filepath}")
            df = df.dropna(how='all')
            df = df.loc[~(df.map(lambda x: str(x).strip()).eq('').all(axis=1))]
            return df

        df_1200 = read_sheet(file_1200_path)
        df_1300 = read_sheet(file_1300_path)

        wb_out     = openpyxl.load_workbook(output_excel_path)
        sheet_name = 'Repair_sheet'
        ws_out     = wb_out[sheet_name] if sheet_name in wb_out.sheetnames else wb_out.create_sheet(sheet_name)

        for row in ws_out.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        for i, row in enumerate(df_1300.itertuples(index=False, name=None), start=3):
            for j, val in enumerate(row, start=1):
                ws_out.cell(row=i, column=j, value=val)
        for i, row in enumerate(df_1200.itertuples(index=False, name=None), start=3):
            for j, val in enumerate(row, start=8):
                ws_out.cell(row=i, column=j, value=val)

        ws_out.merge_cells("A1:G1")
        ws_out.merge_cells("H1:N1")
        ws_out["A1"] = "1300 Repair_Sheet"
        ws_out["H1"] = "1200 Repair_Sheet"

        for col, value in enumerate(["No.", "DUT", "Type", "PAD", "S/L", "X 좌표", "Y 좌표"], start=1):
            ws_out.cell(row=2, column=col, value=value)
        for col, value in enumerate(["No.", "DUT", "정/역", "PAD", "S/L", "X 좌표", "Y 좌표"], start=8):
            ws_out.cell(row=2, column=col, value=value)

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"),  bottom=Side(style="thin"))
        for col in range(1, 15):
            cell = ws_out.cell(row=1, column=col)
            cell.alignment = center_align
            cell.border    = thin_border

        wb_out.save(output_excel_path)
        print(f"'Repair_sheet' 시트 생성 및 데이터 입력 완료.")
    except Exception as e:
        print(f"run_14 오류 발생: {e}")


def run_15():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        df = pd.read_excel(output_excel_path, sheet_name='Repair_sheet', header=1, usecols='A:N')

        if len(sl_values) == 1 and sl_values[0].upper() == "S/L":
            filtered_df = df[df['S/L'].notna() & (df['S/L'].astype(str).str.strip() != '')]
        else:
            filtered_df = df[df['S/L'].isin(sl_values)]

        wb           = load_workbook(output_excel_path)
        filter_sheet = 'Repair_sheet_filter'
        if filter_sheet in wb.sheetnames:
            del wb[filter_sheet]
        ws = wb.create_sheet(filter_sheet)

        ws.merge_cells("A1:G1")
        ws.merge_cells("H1:N1")
        ws["A1"] = "1300 Repair_Sheet"
        ws["H1"] = "1200 Repair_Sheet"

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"),  bottom=Side(style="thin"))
        for col in range(1, 15):
            cell = ws.cell(row=1, column=col)
            cell.alignment = center_align
            cell.border    = thin_border

        for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_excel_path)
        print(f"'Repair_sheet_filter' 시트에 '{', '.join(sl_values)}' 필터 결과 저장 완료.")
    except Exception as e:
        print(f"❌ run_15 오류 발생: {e}")


def run_16():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb         = load_workbook(output_excel_path)
        sheet_name = '1300-1200_Repair'
        ws         = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        ws.merge_cells("A1:C1")
        ws.merge_cells("D1:F1")
        ws["A1"] = "1300 Repair_Sheet"
        ws["D1"] = "1200 Repair_Sheet"
        for col, value in enumerate(["DUT", "PAD", "fail", "정/역", "DUT", "PAD"], start=1):
            ws.cell(row=2, column=col, value=value)

        df_pin = pd.read_excel(output_excel_path, sheet_name='Repair_Pin', usecols=[0, 1, 4], header=0)
        df_pin = df_pin.dropna(how='all')
        for row_idx, (dut, pad, fail) in enumerate(df_pin.itertuples(index=False, name=None), start=3):
            ws.cell(row=row_idx, column=1, value=dut)
            ws.cell(row=row_idx, column=2, value=pad)
            ws.cell(row=row_idx, column=3, value=fail)

        wb.save(output_excel_path)
        print(f"'1300-1200_Repair' 시트에 Repair_Pin 데이터가 추가되었습니다.")
    except Exception as e:
        print(f"run_16 오류 발생: {e}")


def run_17():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb           = load_workbook(output_excel_path)
        repair_sheet = '1300-1200_Repair'
        filter_sheet = 'Repair_sheet_filter'

        if repair_sheet not in wb.sheetnames or filter_sheet not in wb.sheetnames:
            print(f"❌ 필요한 시트가 존재하지 않습니다.")
            return

        ws        = wb[repair_sheet]
        ws_filter = wb[filter_sheet]

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"),  bottom=Side(style="thin"))
        for col in range(1, 7):
            ws.cell(row=1, column=col).alignment = center_align
            ws.cell(row=1, column=col).border    = thin_border
            ws.cell(row=2, column=col).alignment = center_align
            ws.cell(row=2, column=col).border    = thin_border

        filter_data = {}
        for row in ws_filter.iter_rows(min_row=2, values_only=True):
            key   = (row[1], row[3])
            value = (row[9], row[8], row[10])
            filter_data[key] = value

        row_idx = 3
        while True:
            dut = ws.cell(row=row_idx, column=1).value
            pad = ws.cell(row=row_idx, column=2).value
            if dut is None and pad is None:
                break
            match = filter_data.get((dut, pad))
            if match:
                ws.cell(row=row_idx, column=4, value=match[0])
                ws.cell(row=row_idx, column=5, value=match[1])
                ws.cell(row=row_idx, column=6, value=match[2])
            row_idx += 1

        wb.save(output_excel_path)
        print(f"'1300-1200_Repair' 시트에 데이터가 적용되었습니다.")
    except Exception as e:
        print(f"run_17 오류 발생: {e}")


def run_18():
    try:
        output_excel_path = os.path.splitext(st2_file_path)[0] + '.xlsx'
        wb       = load_workbook(output_excel_path)
        src_sheet = '1300-1200_Repair'
        if src_sheet not in wb.sheetnames:
            print(f"❌ '{src_sheet}' 시트가 존재하지 않습니다.")
            return
        ws_src = wb[src_sheet]

        if 'PRINT' in wb.sheetnames:
            del wb['PRINT']
        ws_print = wb.create_sheet('PRINT')
        ws_print.sheet_properties.tabColor = "FFFF00"

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"),  bottom=Side(style="thin"))

        for merge in ["A1:E2","A3:E3","F3:H3","G1:H1","G2:H2",
                      "I3:J3","J1:L1","J2:L2","K3:O3","M1:N1","M2:N2"]:
            ws_print.merge_cells(merge)

        ws_print["A1"] = ", ".join(sl_values) if sl_values else ""
        ws_print["F1"] = "작성자"
        ws_print["F2"] = "Probe"
        ws_print["F3"] = "1300 Repair_Sheet"
        ws_print["I1"] = "제품 No"
        ws_print["I2"] = "설비 No"
        ws_print["I3"] = "1200 Repair_Sheet"
        ws_print["K3"] = "Inspection"
        ws_print["M1"] = "기판"
        ws_print["M2"] = "Fail X Y Z"
        ws_print["O1"] = " / "
        ws_print["O2"] = '=SUM(B5:D1000)&" / "&SUM(B5:B1000)&" / "&SUM(C5:C1000)&" / "&SUM(D5:D1000)'
        ws_print["Q1"] = "Fail"
        ws_print["Q2"] = "X ="
        ws_print["Q3"] = "Y ="
        ws_print["Q4"] = "Z ="
        ws_print["R1"] = '=SUM(B5:D1000)'
        ws_print["R2"] = '=SUM(B5:B1000)'
        ws_print["R3"] = '=SUM(C5:C1000)'
        ws_print["R4"] = '=SUM(D5:D1000)'
        ws_print["T1"] = '전체 Pin'
        ws_print["T2"] = 'Fail'
        ws_print["T3"] = 'Skip'
        ws_print["T4"] = 'Pass'
        ws_print["A3"] = datetime.today().strftime("%y-%m-%d")

        bold_font = Font(bold=True)
        for ref in ["F1","F2","F3","I1","I2","I3","M1","M2","K3"]:
            ws_print[ref].font = bold_font

        headers = ["No","X","Y","Z","정/역","DUT","PAD","빈 PAD","DUT","PAD","냉납","Burnt","눌림","패드X","내용"]
        for col, value in enumerate(headers, start=1):
            cell = ws_print.cell(row=4, column=col, value=value)
            cell.alignment = center_align
            cell.border    = thin_border

        for row in range(1, 4):
            for col in range(1, 16):
                ws_print.cell(row=row, column=col).alignment = center_align
                ws_print.cell(row=row, column=col).border    = thin_border

        thin_black = Side(style="thin", color="000000")
        for row in range(1, 5):
            for cols in (("Q", "R"), ("T", "U")):
                for col in cols:
                    cell = ws_print[f"{col}{row}"]
                    cell.alignment = center_align
                    cell.border    = Border(
                        top=thin_black, bottom=thin_black,
                        left=thin_black  if col == cols[0]  else None,
                        right=thin_black if col == cols[-1] else None
                    )

        def format_pad_list(pads):
            if not pads:
                return ""
            pads   = sorted(set(pads))
            ranges = []
            start  = end = pads[0]
            for p in pads[1:]:
                if p == end + 1:
                    end = p
                else:
                    ranges.append(f"{start}" if start == end else f"{start} ~ {end}")
                    start = end = p
            ranges.append(f"{start}" if start == end else f"{start} ~ {end}")
            return ", ".join(ranges)

        print_row = 5
        pad_B_list, pad_C_list, pad_F_list, pad_G_list = [], [], [], []
        first_dut = first_rev = first_dut2 = None

        for row_idx in range(3, ws_src.max_row + 1):
            dut  = ws_src.cell(row=row_idx, column=1).value
            pad  = ws_src.cell(row=row_idx, column=2).value
            fail = ws_src.cell(row=row_idx, column=3).value
            rev  = ws_src.cell(row=row_idx, column=4).value
            dut2 = ws_src.cell(row=row_idx, column=5).value
            pad2 = ws_src.cell(row=row_idx, column=6).value
            nxt  = row_idx + 1 <= ws_src.max_row
            next_dut  = ws_src.cell(row=row_idx+1, column=1).value if nxt else None
            next_fail = ws_src.cell(row=row_idx+1, column=3).value if nxt else None
            next_rev  = ws_src.cell(row=row_idx+1, column=4).value if nxt else None
            next_dut2 = ws_src.cell(row=row_idx+1, column=5).value if nxt else None

            if first_dut is None:
                first_dut = dut
                first_rev = rev
                first_dut2 = dut2

            if fail == 0:
                pad_C_list.append(pad)
                pad_G_list.append(pad2)
            elif fail == 4:
                pad_B_list.append(pad)
                pad_F_list.append(pad2)

            if (next_dut != dut or next_rev != rev or next_dut2 != dut2) or \
               (next_fail not in [0, 4]) or (next_dut is None):
                ws_print.cell(row=print_row, column=5,  value=first_rev).alignment  = center_align
                ws_print.cell(row=print_row, column=6,  value=first_dut).alignment  = center_align
                ws_print.cell(row=print_row, column=9,  value=first_dut2).alignment = center_align
                ws_print.cell(row=print_row, column=7,  value=format_pad_list(pad_B_list)).alignment = center_align
                ws_print.cell(row=print_row, column=8,  value=format_pad_list(pad_C_list)).alignment = center_align
                ws_print.cell(row=print_row, column=10, value=format_pad_list(pad_F_list)).alignment = center_align
                print_row += 1
                pad_B_list.clear(); pad_C_list.clear()
                pad_F_list.clear(); pad_G_list.clear()
                first_dut = first_rev = first_dut2 = None

        gray_thin  = Side(style="thin", color="BFBFBF")
        black_thin = Side(style="thin", color="000000")
        for row in range(5, print_row):
            for col in range(1, 16):
                cell   = ws_print.cell(row=row, column=col)
                bottom = gray_thin
                left   = black_thin if col in [9, 11] else (gray_thin if col > 1 else Side(border_style=None))
                cell.border = Border(left=left, bottom=bottom)

        fit_c   = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
        ctr_c   = Alignment(horizontal="center", vertical="center")
        for row in range(5, print_row):
            for col in [2, 3, 4, 12, 13, 14, 15]:
                cell = ws_print.cell(row=row, column=col)
                if cell.value not in (None, ""):
                    cell.alignment = ctr_c
            for col in [7, 8, 10, 11]:
                cell = ws_print.cell(row=row, column=col)
                if cell.value not in (None, ""):
                    cell.alignment = fit_c

        e_col_4 = e_col_0 = 0
        if "MAP_PAD" in wb.sheetnames:
            ws_map = wb["MAP_PAD"]
            for r in range(2, ws_map.max_row + 1):
                val = ws_map.cell(r, 5).value
                if val == 4:
                    e_col_4 += 1
                elif val == 0:
                    e_col_0 += 1

        ws_print["U1"] = len([v for v in list_block if isinstance(v, (int, float))])
        ws_print["U2"] = e_col_4
        ws_print["U3"] = e_col_0
        ws_print["U4"] = '=U1-U2-U3'

        print_end_row = print_row - 1
        ws_print.print_area             = f"A1:O{print_end_row}"
        ws_print.print_title_rows       = '1:4'
        ws_print.page_setup.paperSize   = ws_print.PAPERSIZE_A4
        ws_print.page_setup.fitToPage   = True
        ws_print.page_setup.fitToWidth  = 1
        ws_print.page_setup.fitToHeight = 0
        ws_print.page_setup.orientation = ws_print.ORIENTATION_LANDSCAPE
        ws_print.oddFooter.center.text  = "페이지 &P"
        ws_print.page_margins = PageMargins(left=0.25, right=0.25, top=0.4,
                                             bottom=0.6, header=0.3, footer=0.2)

        for col, width in [("A",6),("B",4),("C",4),("D",4),("E",6),
                            ("G",17),("H",17),("J",17),("K",6),("L",6),
                            ("M",6),("N",6),("O",34),("U",8)]:
            ws_print.column_dimensions[col].width = width

        wb.save(output_excel_path)
        print(f"✅ 'PRINT' 시트가 생성되었고, 테두리 및 열 너비가 설정되었습니다.")
    except Exception as e:
        print(f"❌ run_18 오류 발생: {e}")


def main():
    runs = [run_01, run_02, run_03, run_04, run_05, run_06,
            run_07, run_08, run_09, run_10, run_11, run_12,
            run_13, run_14, run_15, run_16, run_17, run_18]
    try:
        for i, fn in enumerate(runs, 1):
            if stop_event.is_set():
                print("⛔ 실행이 중단되었습니다.")
                return
            print(f"{i:02d} 실행 중")
            fn()
        print("모든 스크립트 실행 완료!")
    except Exception as e:
        with open("error_log.txt", "w", encoding="utf-8") as f:
            traceback.print_exc(file=f)
        print("에러가 발생했습니다. error_log.txt를 확인하세요.")


# ============================================================
# XYZ Fail — 처리 함수들
# ============================================================

def ftp_list_001(ftp_folder: str, setting_dt: datetime, st2_modified_dt: datetime, out_filename="FTP_List.csv"):
    ftp_path = Path(ftp_folder)
    if not ftp_path.is_dir():
        raise ValueError(f"FTP 폴더가 올바르지 않습니다: {ftp_folder}")

    out_path  = ftp_path.parent / out_filename
    csv_files = sorted(ftp_path.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"FTP 폴더 안에 CSV 파일이 없습니다: {ftp_folder}")

    dfs = []
    base_cols = None
    for fp in csv_files:
        df = pd.read_csv(fp, encoding="utf-8-sig", dtype=str)
        if base_cols is None:
            base_cols = list(df.columns)
        extra_cols = [c for c in df.columns if c not in base_cols]
        df = df[base_cols + extra_cols]
        df.insert(0, "_source_file", fp.name)
        dfs.append(df)

    merged = pd.concat(dfs, ignore_index=True)
    cols   = list(merged.columns)
    if len(cols) < 4:
        raise ValueError("컬럼 수가 4개 미만이라 B/C/D 기준 작업이 불가능합니다.")

    b_col   = cols[1]
    cd_cols = [cols[2], cols[3]]
    merged["_B_DT_"] = pd.to_datetime(
        merged[b_col].astype(str).str.replace(r"\s+", " ", regex=True).str.strip(),
        errors="coerce"
    )
    merged = merged[merged["_B_DT_"].notna()].copy()
    merged = merged[(merged["_B_DT_"] >= setting_dt) & (merged["_B_DT_"] <= st2_modified_dt)].copy()
    merged = merged.sort_values("_B_DT_", ascending=False)
    merged = merged.drop_duplicates(subset=cd_cols, keep="first")
    merged = merged.sort_values("_B_DT_", ascending=True).reset_index(drop=True)
    merged.drop(columns=["_B_DT_"], inplace=True)
    merged.to_csv(out_path, index=False, encoding="utf-8-sig")

    return merged.to_dict(orient="records"), str(out_path), [str(p) for p in csv_files]


def xy_po_001(print_excel_path: str, FTP_List: list, spec_x: float, spec_y: float):
    wb = load_workbook(print_excel_path)
    if "1300-1200_Repair" not in wb.sheetnames:
        raise ValueError("'1300-1200_Repair' 시트가 없습니다.")

    if "XY_PO" in wb.sheetnames:
        del wb["XY_PO"]
    ws_new = wb.copy_worksheet(wb["1300-1200_Repair"])
    ws_new.title = "XY_PO"

    rows_to_delete = []
    last_row = ws_new.max_row
    for row in range(3, last_row + 1):
        val = ws_new.cell(row=row, column=3).value
        try:
            if int(float(str(val).strip())) != 4:
                rows_to_delete.append(row)
        except Exception:
            rows_to_delete.append(row)
    for r in reversed(rows_to_delete):
        ws_new.delete_rows(r)

    if not FTP_List:
        wb.save(print_excel_path)
        return

    ftp_cols = list(FTP_List[0].keys())
    if len(ftp_cols) < 8:
        raise ValueError("FTP_List 컬럼이 8개 미만입니다.")

    ftp_C, ftp_D, ftp_G, ftp_H = ftp_cols[2], ftp_cols[3], ftp_cols[6], ftp_cols[7]

    def norm(v):
        return "" if v is None else str(v).strip()

    def to_number(v):
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        try:
            f = float(s)
            return int(f) if f.is_integer() else f
        except Exception:
            return v

    ftp_map = {
        (norm(r.get(ftp_C)), norm(r.get(ftp_D))): (r.get(ftp_C), r.get(ftp_D), r.get(ftp_G), r.get(ftp_H))
        for r in FTP_List
    }

    for row in range(3, ws_new.max_row + 1):
        key = (norm(ws_new.cell(row=row, column=1).value), norm(ws_new.cell(row=row, column=2).value))
        if key in ftp_map:
            c_val, d_val, g_val, h_val = ftp_map[key]
            ws_new.cell(row=row, column=7).value  = to_number(c_val)
            ws_new.cell(row=row, column=8).value  = to_number(d_val)
            ws_new.cell(row=row, column=9).value  = to_number(g_val)
            ws_new.cell(row=row, column=10).value = to_number(h_val)

    judge_align = Alignment(horizontal="center", vertical="center")
    judge_font  = Font(bold=True)

    def to_float(v):
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).strip())
        except Exception:
            return None

    for row in range(3, ws_new.max_row + 1):
        x = to_float(ws_new.cell(row=row, column=9).value)
        y = to_float(ws_new.cell(row=row, column=10).value)
        if x is None or y is None:
            result = "-"
        else:
            ax, ay = abs(x), abs(y)
            result = "Y" if (ax > spec_x and ay > spec_y) else ("X" if ax > spec_x else ("Y" if ay > spec_y else "-"))
        cell = ws_new.cell(row=row, column=11)
        cell.value     = result
        cell.alignment = judge_align
        cell.font      = judge_font

    header_align = Alignment(horizontal="center", vertical="center")
    header_font  = Font(bold=True)
    for col, text in {7:"DUT", 8:"PAD", 9:"X", 10:"Y", 11:"판정"}.items():
        cell = ws_new.cell(row=2, column=col)
        cell.value     = text
        cell.alignment = header_align
        cell.font      = header_font

    if "PRINT" in wb.sheetnames:
        ws_print = wb["PRINT"]
        xy_to_k = {}
        for r in range(3, ws_new.max_row + 1):
            a = ws_new.cell(r, 1).value
            b = ws_new.cell(r, 2).value
            k = ws_new.cell(r, 11).value
            key = (norm(a), norm(b))
            if key != ("", ""):
                xy_to_k[key] = k

        for r in range(3, ws_print.max_row + 1):
            d = ws_print.cell(r, 6).value
            e = ws_print.cell(r, 7).value
            if d is None or e is None:
                continue
            d_norm  = norm(d)
            results = []
            for part in str(e).split(","):
                part = part.strip()
                if not part:
                    continue
                if "~" in part:
                    try:
                        s_p, e_p = [int(x.strip()) for x in part.split("~")]
                        for p in range(s_p, e_p + 1):
                            val = xy_to_k.get((d_norm, str(p)))
                            if val is not None:
                                results.append("Z" if val == "-" else str(val))
                    except Exception:
                        pass
                    continue
                val = xy_to_k.get((d_norm, norm(part)))
                if val is not None:
                    results.append("Z" if val == "-" else str(val))

            if results:
                if (n := results.count("X")):  ws_print.cell(r, 2).value = n
                if (n := results.count("Y")):  ws_print.cell(r, 3).value = n
                if (n := results.count("Z")):  ws_print.cell(r, 4).value = n

    wb.save(print_excel_path)


# ============================================================
# 공통 색상 / 스타일 상수
# ============================================================
BG        = "#1f1f1f"   # 메인 배경
BG2       = "#2d2d2d"   # 입력창 / 버튼 배경
FG        = "#ffffff"   # 기본 텍스트
FG_DIM    = "#888888"   # 보조 텍스트 (파일명)
ACCENT    = "#d35400"   # 오렌지 (선택·강조)
ACCENT_HI = "#ff8c00"   # 오렌지 호버
FONT      = ("Malgun Gothic", 10)
FONT_BOLD = ("Malgun Gothic", 10, "bold")
FONT_SM   = ("Malgun Gothic", 9)


def _entry_kw(width=100):
    return dict(bg=BG2, fg=FG, insertbackground=FG,
                font=FONT, width=width, relief="flat")

def _lbl_btn_kw():
    return dict(bg=BG, fg=FG, font=FONT_BOLD, relief="flat",
                bd=0, cursor="hand2", highlightthickness=0,
                activebackground=BG, activeforeground=ACCENT_HI)

def _run_btn_kw():
    return dict(bg=BG2, fg=FG, font=FONT_BOLD,
                width=10, relief="raised", bd=2,
                activebackground=ACCENT, activeforeground=FG)


# ============================================================
# Tab 1: Repair PIN
# ============================================================
class FailPinMapTab(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG, padx=16, pady=12)
        self.st2_var   = tk.StringVar()
        self.csv_var   = tk.StringVar()
        self.r1200_var = tk.StringVar()
        self.r1300_var = tk.StringVar()
        self.sl_var    = tk.StringVar()
        self._fname_lbls = {}
        self._worker_thread = None
        self._build()

    def _build(self):
        self.columnconfigure(1, weight=1)

        # Row 0: S/L 입력
        tk.Button(self, text="S/L 입력", **_lbl_btn_kw(),
                  command=lambda: None).grid(row=0, column=0, sticky="e", padx=(0, 14), pady=6)
        sl_frame = tk.Frame(self, bg=BG)
        sl_frame.grid(row=0, column=1, sticky="w", pady=6)
        self.sl_combo = ttk.Combobox(sl_frame, textvariable=self.sl_var,
                                     style="Dark.TCombobox", width=22)
        self.sl_combo.pack(side="left")
        self.lbl_col_count = tk.Label(sl_frame, text="", bg=BG, fg=FG_DIM, font=FONT_SM)
        self.lbl_col_count.pack(side="left", padx=(6, 0))

        # Rows 1-4: 파일 선택
        fields = [
            ("ST2 파일",    self.st2_var,   "st2"),
            ("CSV 파일",    self.csv_var,   "csv"),
            ("Repair 1200", self.r1200_var, "1200"),
            ("Repair 1300", self.r1300_var, "1300"),
        ]
        self._picker_btns = []
        for i, (label, var, kind) in enumerate(fields, start=1):
            btn = tk.Button(self, text=f"{label}  ›", **_lbl_btn_kw(),
                            command=lambda k=kind: self._pick(k))
            btn.grid(row=i, column=0, sticky="ne", padx=(0, 14), pady=(6, 0))
            self._picker_btns.append(btn)

            inner = tk.Frame(self, bg=BG)
            inner.grid(row=i, column=1, sticky="w", pady=(6, 0))
            tk.Entry(inner, textvariable=var, **_entry_kw()).pack(anchor="w")
            lbl = tk.Label(inner, text="", bg=BG, fg=FG_DIM, font=FONT_SM)
            lbl.pack(anchor="w", pady=(2, 0))
            self._fname_lbls[kind] = (var, lbl)

        # Row 5: 실행 / 정지 버튼
        btn_row = tk.Frame(self, bg=BG)
        btn_row.grid(row=5, column=1, sticky="e", pady=(14, 0))
        self.btn_stop = tk.Button(btn_row, text="정지",
            bg="#5a1a1a", fg=FG, font=FONT_BOLD, width=6, relief="raised", bd=2,
            activebackground="#8b0000", activeforeground=FG,
            state="disabled", command=self._stop)
        self.btn_stop.pack(side="right", padx=(6, 0))
        self.btn_run = tk.Button(btn_row, text="실행", **_run_btn_kw(), command=self._run)
        self.btn_run.pack(side="right")

    def _pick(self, kind):
        specs = {
            "st2":  ("ST2 파일 선택",         [("ST2 파일","*.st2"),  ("모든 파일","*.*")], False),
            "csv":  ("CSV 파일 선택",         [("CSV 파일","*.csv"),  ("모든 파일","*.*")], False),
            "1200": ("1200 Repair_sheet 선택",[("Excel 파일","*.xlsx;*.xls"),("모든 파일","*.*")], False),
            "1300": ("1300 Repair_sheet 선택",[("Excel 파일","*.xlsx;*.xls"),("모든 파일","*.*")], False),
        }
        title, ftypes, _ = specs[kind]
        p = filedialog.askopenfilename(title=title, filetypes=ftypes)
        var, lbl = self._fname_lbls[kind]
        if p:
            var.set(p)
            lbl.config(text=os.path.basename(p), fg=ACCENT)
            if kind == "1300":
                self._update_sl_from_1300(p)
        else:
            lbl.config(text="", fg=FG_DIM)

    def _update_sl_from_1300(self, path):
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext in ['.xlsx', '.xls']:
                df_tmp = pd.read_excel(path, header=None)
            else:
                try:
                    df_tmp = pd.read_csv(path, encoding='cp949', header=None)
                except UnicodeDecodeError:
                    df_tmp = pd.read_csv(path, encoding='utf-8', header=None)

            def is_numeric(v):
                try:
                    float(v)
                    return True
                except (ValueError, TypeError):
                    return False

            num_cols = df_tmp.shape[1]
            self.lbl_col_count.config(text=f"{num_cols}열")
            if num_cols == 7:
                raw_vals = df_tmp.iloc[:, 4].dropna().astype(str).str.strip()
                unique_vals = sorted(set(
                    v for v in raw_vals
                    if v and v.lower() != 's/l' and not is_numeric(v)
                ))
                self.sl_combo['values'] = unique_vals
                self.sl_var.set(unique_vals[0] if unique_vals else 'S/L')
            elif num_cols == 6:
                self.sl_combo['values'] = ['S/L']
                self.sl_var.set('S/L')
            else:
                messagebox.showwarning("열 수 오류",
                    f"1300 파일이 {num_cols}열입니다.\n7열로 수정해주세요.")
                self.sl_combo['values'] = []
                self.sl_var.set('')
        except Exception as e:
            print(f"1300 파일 E열 읽기 오류: {e}")

    def _run(self):
        global st2_file_path, csv_file_path, file_1200_path, file_1300_path, sl_values, stop_event
        st2_file_path  = self.st2_var.get().strip()
        csv_file_path  = self.csv_var.get().strip()
        file_1200_path = self.r1200_var.get().strip()
        file_1300_path = self.r1300_var.get().strip()
        sl_values      = parse_sl_input(self.sl_var.get().strip())

        print(f"S/L 값: {sl_values}")
        print(f"ST2 파일: {st2_file_path}")
        print(f"CSV 파일: {csv_file_path}")

        stop_event.clear()
        for btn in self._picker_btns:
            btn.config(state="disabled")
        self.sl_combo.config(state="disabled")
        self.btn_run.config(text="처리 중...", state="disabled",
                            bg=ACCENT, activebackground=ACCENT)
        self.btn_stop.config(state="normal")
        self._worker_thread = threading.Thread(target=self._worker, daemon=True)
        self._worker_thread.start()

    def _stop(self):
        if self._worker_thread and self._worker_thread.is_alive():
            ctypes.pythonapi.PyThreadState_SetAsyncExc(
                ctypes.c_ulong(self._worker_thread.ident),
                ctypes.py_object(_StopExecution)
            )
        self.btn_stop.config(state="disabled", text="정지 중...")

    def _worker(self):
        try:
            main()
        except _StopExecution:
            print("⛔ 실행이 중단되었습니다.")
        self.after(0, self._restore_ui)

    def _restore_ui(self):
        self.btn_run.config(text="실행", state="normal", **_run_btn_kw())
        self.btn_stop.config(state="disabled", text="정지")
        self.sl_combo.config(state="normal")
        for btn in self._picker_btns:
            btn.config(state="normal")


# ============================================================
# Tab 2: XYZ Fail
# ============================================================
class XYZFailTab(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG, padx=16, pady=12)
        now = datetime.now()
        self.year_var   = tk.IntVar(value=now.year)
        self.mon_var    = tk.IntVar(value=now.month)
        self.day_var    = tk.IntVar(value=now.day)
        self.hour_var   = tk.IntVar(value=now.hour)
        self.min_var    = tk.IntVar(value=now.minute)
        self.print_path = tk.StringVar()
        self.st2_path   = tk.StringVar()
        self.ftp_dir    = tk.StringVar()
        self.spec_x     = tk.StringVar(value="3")
        self.spec_y     = tk.StringVar(value="3")
        self.status_var = tk.StringVar(value="")
        self._fname_lbls = {}
        self._build()

    def _build(self):
        self.columnconfigure(1, weight=1)

        # Row 0: 세팅 시간
        tk.Label(self, text="세팅 시간  ›", bg=BG, fg=FG,
                 font=FONT_BOLD).grid(row=0, column=0, sticky="e", padx=(0, 14), pady=6)
        tf = tk.Frame(self, bg=BG)
        tf.grid(row=0, column=1, sticky="w", pady=6)
        self.cmb_year = self._combo(tf, self.year_var, list(range(2000, 2101)), "년", 6)
        self.cmb_mon  = self._combo(tf, self.mon_var,  list(range(1, 13)),      "월", 4)
        self.cmb_day  = self._combo(tf, self.day_var,  [],                      "일", 4)
        self.cmb_hour = self._combo(tf, self.hour_var, list(range(0, 24)),      "시", 4)
        self.cmb_min  = self._combo(tf, self.min_var,  list(range(0, 60)),      "분", 4)
        self._refresh_days()
        self.cmb_year.bind("<<ComboboxSelected>>", lambda e: self._refresh_days())
        self.cmb_mon.bind("<<ComboboxSelected>>",  lambda e: self._refresh_days())

        # Rows 1-3: 파일 선택
        file_fields = [
            ("Print 파일", self.print_path, self._pick_print),
            ("ST2 파일",   self.st2_path,   self._pick_st2),
            ("FTP 폴더",   self.ftp_dir,    self._pick_ftp),
        ]
        for i, (label, var, cmd) in enumerate(file_fields, start=1):
            tk.Button(self, text=f"{label}  ›", **_lbl_btn_kw(),
                      command=cmd).grid(row=i, column=0, sticky="ne", padx=(0, 14), pady=(6, 0))
            inner = tk.Frame(self, bg=BG)
            inner.grid(row=i, column=1, sticky="w", pady=(6, 0))
            tk.Entry(inner, textvariable=var, **_entry_kw()).pack(anchor="w")
            lbl = tk.Label(inner, text="", bg=BG, fg=FG_DIM, font=FONT_SM)
            lbl.pack(anchor="w", pady=(2, 0))
            self._fname_lbls[label] = (var, lbl)

        # Row 4: Spec X / Y
        tk.Label(self, text="Spec X / Y  ›", bg=BG, fg=FG,
                 font=FONT_BOLD).grid(row=4, column=0, sticky="e", padx=(0, 14), pady=6)
        spec_frame = tk.Frame(self, bg=BG)
        spec_frame.grid(row=4, column=1, sticky="w", pady=6)
        tk.Label(spec_frame, text="X >", bg=BG, fg=FG_DIM, font=FONT).pack(side="left")
        tk.Entry(spec_frame, textvariable=self.spec_x,
                 **_entry_kw(width=6), justify="center").pack(side="left", padx=(4, 14))
        tk.Label(spec_frame, text="Y >", bg=BG, fg=FG_DIM, font=FONT).pack(side="left")
        tk.Entry(spec_frame, textvariable=self.spec_y,
                 **_entry_kw(width=6), justify="center").pack(side="left", padx=(4, 0))

        # Row 5: 실행 버튼 + 상태 라벨
        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.grid(row=5, column=1, sticky="e", pady=(10, 0))
        self.btn_run = tk.Button(btn_frame, text="실행", **_run_btn_kw(), command=self._run)
        self.btn_run.pack(side="right")
        tk.Label(btn_frame, textvariable=self.status_var,
                 bg=BG, fg=ACCENT, font=FONT_SM).pack(side="right", padx=14)

    def _combo(self, parent, var, values, suffix, width):
        cmb = ttk.Combobox(parent, style="Dark.TCombobox",
                           state="readonly", width=width, justify="center")
        cmb.pack(side="left")
        tk.Label(parent, text=suffix, bg=BG, fg=FG_DIM,
                 font=FONT, padx=4).pack(side="left")
        cmb["values"] = [str(v) for v in values]
        target = str(var.get())
        vals   = list(cmb["values"])
        cmb.set(target if target in vals else (vals[0] if vals else ""))
        def on_select(_e=None):
            try: var.set(int(cmb.get()))
            except Exception: pass
        cmb.bind("<<ComboboxSelected>>", on_select)
        return cmb

    def _refresh_days(self):
        last_day = calendar.monthrange(int(self.year_var.get()), int(self.mon_var.get()))[1]
        day_vals = [str(d) for d in range(1, last_day + 1)]
        self.cmb_day["values"] = day_vals
        cur = str(self.day_var.get())
        self.cmb_day.set(cur if cur in day_vals else day_vals[-1])

    def _set_fname(self, key, path):
        var, lbl = self._fname_lbls[key]
        if path:
            var.set(path)
            lbl.config(text=os.path.basename(path), fg=ACCENT)
        else:
            lbl.config(text="", fg=FG_DIM)

    def _pick_print(self):
        p = filedialog.askopenfilename(title="Print 파일 선택",
                filetypes=[("Excel 파일","*.xlsx;*.xlsm;*.xls"),("모든 파일","*.*")])
        self._set_fname("Print 파일", p)

    def _pick_st2(self):
        p = filedialog.askopenfilename(title="ST2 파일 선택",
                filetypes=[("ST2 파일","*.st2"),("모든 파일","*.*")])
        if not p: return
        self._set_fname("ST2 파일", p)
        dt = datetime.fromtimestamp(os.path.getmtime(p))
        self.status_var.set(f"마감 시간: {dt:%Y-%m-%d %H:%M:%S}")
        self._refresh_days()

    def _pick_ftp(self):
        p = filedialog.askdirectory(title="FTP 폴더 선택")
        self._set_fname("FTP 폴더", p)

    def _run(self):
        self.btn_run.config(text="처리 중...", state="disabled",
                            bg=ACCENT, activebackground=ACCENT)
        self.status_var.set("작업 중...")
        threading.Thread(target=self._run_worker, daemon=True).start()

    def _run_worker(self):
        def _restore():
            self.btn_run.config(text="실행", state="normal", **_run_btn_kw())
        try:
            try:
                setting_dt = datetime(int(self.cmb_year.get()), int(self.cmb_mon.get()),
                                      int(self.cmb_day.get()),  int(self.cmb_hour.get()),
                                      int(self.cmb_min.get()))
            except Exception:
                self.after(0, lambda: messagebox.showerror("오류", "세팅 시간이 올바르지 않습니다."))
                return

            print_file = self._fname_lbls["Print 파일"][0].get().strip()
            st2_file   = self._fname_lbls["ST2 파일"][0].get().strip()
            ftp_folder = self._fname_lbls["FTP 폴더"][0].get().strip()

            if not print_file or not os.path.isfile(print_file):
                self.after(0, lambda: messagebox.showerror("입력 오류", "Print 파일을 선택해주세요."))
                return
            if not st2_file or not os.path.isfile(st2_file):
                self.after(0, lambda: messagebox.showerror("입력 오류", "ST2 파일을 선택해주세요."))
                return
            if not ftp_folder or not os.path.isdir(ftp_folder):
                self.after(0, lambda: messagebox.showerror("입력 오류", "FTP 폴더를 선택해주세요."))
                return

            st2_modified_dt = datetime.fromtimestamp(os.path.getmtime(st2_file))

            try:
                FTP_List, saved_path, merged_files = ftp_list_001(ftp_folder, setting_dt, st2_modified_dt)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("병합 오류", str(e)))
                return

            try:
                spec_x = float(self.spec_x.get().strip())
                spec_y = float(self.spec_y.get().strip())
            except Exception:
                self.after(0, lambda: messagebox.showerror("입력 오류", "Spec X/Y 값이 올바르지 않습니다."))
                return

            try:
                xy_po_001(print_file, FTP_List, spec_x, spec_y)
            except Exception as e:
                msg = str(e)
                self.after(0, lambda m=msg: messagebox.showerror("엑셀 처리 오류", m))
                return

            def show_done():
                print(f"✅ XYZ Fail 완료 — 총 {len(FTP_List)}행, {saved_path}")
                messagebox.showinfo("완료",
                    f"세팅 시간: {setting_dt:%Y-%m-%d %H:%M}\n"
                    f"마감 시간: {st2_modified_dt:%Y-%m-%d %H:%M}\n"
                    f"병합 파일 수: {len(merged_files)}\n"
                    f"총 행 수: {len(FTP_List)}\n"
                    f"저장 위치: {saved_path}")
            self.after(0, show_done)

        except Exception as e:
            self.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.after(0, _restore)
            self.after(0, lambda: self.status_var.set("완료"))


# ============================================================
# 메인 앱
# ============================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MAP to Fail")
        self.configure(bg=BG)
        self.resizable(True, False)
        self._dark_titlebar()
        self._setup_ttk_style()
        self._build()

    def _setup_ttk_style(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Dark.TCombobox",
            fieldbackground=BG2, background=BG2, foreground=FG,
            selectbackground=ACCENT, selectforeground=FG,
            arrowcolor=FG_DIM, bordercolor="#444444",
            lightcolor=BG2, darkcolor=BG2, relief="flat", padding=2)
        style.map("Dark.TCombobox",
            fieldbackground=[("readonly", BG2), ("disabled", BG)],
            foreground=[("readonly", FG), ("disabled", FG_DIM)],
            selectbackground=[("readonly", ACCENT)],
            selectforeground=[("readonly", FG)],
            background=[("readonly", BG2), ("active", BG2)])
        self.option_add("*TCombobox*Listbox.background", BG2)
        self.option_add("*TCombobox*Listbox.foreground", FG)
        self.option_add("*TCombobox*Listbox.selectBackground", ACCENT)
        self.option_add("*TCombobox*Listbox.selectForeground", FG)

    def _dark_titlebar(self):
        self.update_idletasks()
        try:
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
            val  = ctypes.c_int(2)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, ctypes.byref(val), ctypes.sizeof(val))
            val  = ctypes.c_int(1)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 19, ctypes.byref(val), ctypes.sizeof(val))
        except Exception:
            pass

    def _build(self):
        # ── 탭 바 ──────────────────────────────────────────
        tab_bar = tk.Frame(self, bg=BG)
        tab_bar.pack(fill="x")

        self._tab_labels    = []
        self._tab_indicators = []

        content = tk.Frame(self, bg=BG)
        content.pack(fill="both", expand=True)
        content.columnconfigure(0, weight=1)
        content.rowconfigure(0, weight=1)

        self._tab1 = FailPinMapTab(content)
        self._tab2 = XYZFailTab(content)
        self._tab_frames = [self._tab1, self._tab2]

        for frame in self._tab_frames:
            frame.grid(row=0, column=0, sticky="nsew")

        for i, title in enumerate(["  Repair PIN  ", "  XYZ Fail  "]):
            col = tk.Frame(tab_bar, bg=BG, cursor="hand2")
            col.pack(side="left")

            lbl = tk.Label(col, text=title, bg=BG, fg=FG_DIM,
                           font=FONT, pady=9, cursor="hand2")
            lbl.pack()

            ind = tk.Frame(col, height=2, bg=BG)
            ind.pack(fill="x")

            self._tab_labels.append(lbl)
            self._tab_indicators.append(ind)

            for w in (col, lbl):
                w.bind("<Button-1>", lambda e, idx=i: self._switch(idx))

        # 구분선
        tk.Frame(self, height=1, bg="#3a3a3a").pack(fill="x")

        # ── 콘솔 ───────────────────────────────────────────
        con = tk.Frame(self, bg="#141414")
        con.pack(fill="x", side="bottom")

        hdr = tk.Frame(con, bg="#1a1a1a")
        hdr.pack(fill="x")
        tk.Label(hdr, text="  출력", bg="#1a1a1a", fg="#555555",
                 font=("Consolas", 8), anchor="w").pack(side="left")
        tk.Button(hdr, text="지우기", bg="#1a1a1a", fg="#555555",
                  font=("Consolas", 8), relief="flat", cursor="hand2",
                  bd=0, highlightthickness=0,
                  activebackground="#1a1a1a", activeforeground=ACCENT_HI,
                  command=self._clear).pack(side="right", padx=6, pady=1)

        self._console = tk.Text(
            con, height=10,
            bg="#141414", fg="#cccccc",
            font=("Consolas", 9),
            state="disabled", wrap="word",
            borderwidth=0, relief="flat",
        )
        sb = tk.Scrollbar(con, bg="#1a1a1a", troughcolor="#141414",
                          activebackground=ACCENT, relief="flat", bd=0,
                          command=self._console.yview)
        self._console.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._console.pack(fill="x")

        sys.stdout = _ConsoleRedirect(self._console)
        self._switch(0)

    def _switch(self, idx):
        for i, (lbl, ind, frame) in enumerate(
                zip(self._tab_labels, self._tab_indicators, self._tab_frames)):
            if i == idx:
                lbl.config(fg=FG, font=FONT_BOLD)
                ind.config(bg=ACCENT)
                frame.tkraise()
            else:
                lbl.config(fg=FG_DIM, font=FONT)
                ind.config(bg=BG)

    def _clear(self):
        self._console.configure(state="normal")
        self._console.delete("1.0", "end")
        self._console.configure(state="disabled")

    def destroy(self):
        sys.stdout = sys.__stdout__
        super().destroy()


if __name__ == "__main__":
    App().mainloop()
