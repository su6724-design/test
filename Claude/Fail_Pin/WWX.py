import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import re
import threading
from openpyxl.styles import PatternFill
import win32com.client
import pythoncom

def select_file(entry_widget, filename_label, extensions):
    """확장자 필터가 적용된 파일 선택 창을 열고 선택된 경로를 엔트리에 설정합니다."""
    # 입력받은 모든 확장자를 하나의 그룹으로 묶습니다.
    patterns = " ".join([f"*{ext}" for ext in extensions])
    filetypes = [("Excel files", patterns), ("All files", "*.*")]
    filename = filedialog.askopenfilename(title="Select file", filetypes=filetypes)
    # 파일을 선택했든 취소했든 먼저 기존 내용을 지웁니다.
    entry_widget.delete(0, tk.END)
    # 새로운 파일이 선택된 경우에만 경로를 삽입하고 파일명을 표시합니다.
    if filename:
        entry_widget.insert(0, filename)
        filename_label.config(text=os.path.basename(filename), fg="#ff8c00") # 파일명만 표시 (주황색 강조)
    else:
        filename_label.config(text="", fg="#888888") # 선택 안함 시 빈칸

def process_files(entries, sort_type_var, stop_event):
    def check_stop():
        if stop_event.is_set():
            raise InterruptedError()
    
    # 필수 파일 선택 여부 확인
    if not entries[0].get() or not entries[1].get():
        messagebox.showerror("Error", "Repair 1300 및 Repair 1200 파일은 반드시 선택해야 합니다.")
        return
    # 파일 경로 가져오기
    rep1300_path = entries[0].get()
    rep1200_path = entries[1].get()
    CM_path = entries[2].get()
    data760_path = entries[3].get()

    # ---------- Repair 1300 ----------
    check_stop()
    df1300 = pd.read_excel(rep1300_path, header=None)
    check_stop()
    col_cnt = df1300.shape[1]
    if col_cnt == 6:
        header1300 = ["P 1300", "DUT", "Type", "PAD", "X좌표", "Y좌표"]
    elif col_cnt == 7:
        header1300 = ["P 1300", "DUT", "Type", "PAD", "S/L", "X좌표", "Y좌표"]
    elif col_cnt == 8:
        header1300 = ["P 1300", "DUT", "Type", "정/역", "PAD", "S/L", "X좌표", "Y좌표"]
    else:
        header1300 = [f"col{i}" for i in range(1, col_cnt + 1)]
    df1300.columns = header1300

    # ---------- Repair 1200 ----------
    check_stop()
    df1200 = pd.read_excel(rep1200_path, header=None)
    check_stop()
    # 고정 헤더 설정 (실제 데이터 열 개수에 맞춰 조정)
    header1200 = ["S 1200", "S DUT", "S 정/역", "S PAD", "S S/L", "X 좌표", "Y 좌표"]
    df1200.columns = header1200[: df1200.shape[1]]

    # ---------- WWX DATA ----------
    df_CM = None
    df_fail = None
    if CM_path:
        xl = pd.ExcelFile(CM_path)
        sheet_names = xl.sheet_names
        # 데이터가 있는 시작 지점을 찾기 위해 먼저 읽어옵니다.
        temp_df = pd.read_excel(CM_path, sheet_name=sheet_names[0], header=None)
        
        # 1열(index 0)에 데이터가 나타나는 첫 번째 행을 헤더 위치로 추정합니다.
        skip_rows = 0
        for i, val in enumerate(temp_df.iloc[:, 0]):
            if pd.notna(val) and str(val).strip() != "":
                skip_rows = i
                break
        
        check_stop()
        # 찾은 행을 헤더로 하여 데이터를 다시 로드합니다.
        df_CM = pd.read_excel(CM_path, sheet_name=sheet_names[0], header=skip_rows)
        check_stop()

        # ---------- WWX Fail (Second Sheet) ----------
        if len(sheet_names) > 1:
            # 2행(index 1)이 헤더, 데이터는 그 이후
            # B열(index 1)부터 최대 M열(index 12)까지만 가져오도록 안전하게 처리
            df_fail_all = pd.read_excel(CM_path, sheet_name=sheet_names[1], header=1)
            df_fail = df_fail_all.iloc[:, 1:13] # B(1) ~ M(12) 범위, 열이 부족해도 에러 발생 안 함
            # 빈 행 제거 (데이터 시작 전 빈 행이 있을 수 있으므로)
            df_fail = df_fail.dropna(how='all').reset_index(drop=True)

    # ---------- 760 Data ----------
    data760_list = []
    if data760_path:
        check_stop()
        # 12번째 행이 헤더(인덱스 11), 데이터는 13번째 행부터 시작
        df760 = pd.read_excel(data760_path, header=11)
        check_stop()
        # Repair 1300의 첫 번째 열 값과 일치하는 행만 필터링
        mask = df760.iloc[:, 0].isin(df1300.iloc[:, 0])
        df760 = df760[mask]
        data760_list = df760.values.tolist()

    # ---------- 리페어 데이터 합치기 ----------
    # 리페어 결과를 가로(열) 방향으로 병합 (빈 열로 구분)
    blank_col = pd.DataFrame({"": [""] * df1300.shape[0]})
    Repair_df = pd.concat([df1300.reset_index(drop=True), blank_col, df1200.reset_index(drop=True)], axis=1)

    # 가로(horizontal) 방식 정렬 및 엑셀 저장
    try:
        # 1. Repair_df_sort 정렬 (가로/세로 공통: S DUT 오름 -> X 좌표 오름 -> Y 좌표 내림 -> S 1200 오름)
        sort_cols_rep = ["S DUT", "X 좌표", "Y 좌표", "S 1200"]
        asc_list_rep = [True, True, False, True]
        Repair_df_sort = Repair_df.copy()
        for col in sort_cols_rep:
            if col in Repair_df_sort.columns:
                Repair_df_sort[col] = pd.to_numeric(Repair_df_sort[col], errors='coerce')
        Repair_df_sort = Repair_df_sort.sort_values(by=sort_cols_rep, ascending=asc_list_rep)
        check_stop()
        
        # 2. df_CM_sort 정렬 (방식에 따라 기준 다름)
        df_CM_sort = pd.DataFrame()
        if df_CM is not None:
            # 가로 배치: STC NO.(오름), X Reference(내림), Y Reference(내림), Probe ID(오름)
            sort_cols_CM = ["STC NO.", "X Reference", "Y Reference", "Probe ID"]
            asc_list_CM = [True, False, False, True]
            
            df_CM_sort = df_CM.copy()
            
            # 존재하는 컬럼 확인 및 숫자 변환
            actual_CM_cols = [c for c in sort_cols_CM if c in df_CM_sort.columns]
            actual_CM_asc = [asc_list_CM[sort_cols_CM.index(c)] for c in actual_CM_cols]
            
            for col in actual_CM_cols:
                df_CM_sort[col] = pd.to_numeric(df_CM_sort[col], errors='coerce')
            
            if actual_CM_cols:
                df_CM_sort = df_CM_sort.sort_values(by=actual_CM_cols, ascending=actual_CM_asc)
            check_stop()
        
        # 3. df_fail 매칭 (df_CM_sort의 Probe ID 열 기준)
        df_fail_matched = pd.DataFrame()
        if df_fail is not None and not df_CM_sort.empty:
            if 'Probe ID' in df_CM_sort.columns and len(df_fail.columns) > 1:
                fail_key_col = df_fail.columns[1]
                
                # 조인 키 이름이 같아서 같이 삭제되는 것을 막기 위해 df_fail 쪽 컬럼명에 '.' 추가
                if fail_key_col == 'Probe ID':
                    df_fail = df_fail.rename(columns={fail_key_col: fail_key_col + '.'})
                    fail_key_col = fail_key_col + '.'
                    
                # df_CM_sort의 Probe ID 순서에 맞게 df_fail 데이터를 병합
                df_fail_matched = pd.merge(
                    df_CM_sort[['Probe ID']], 
                    df_fail, 
                    left_on='Probe ID', 
                    right_on=fail_key_col, 
                    how='left'
                ).drop(columns=['Probe ID']) # 왼쪽 조인 키(df_CM_sort의 Probe ID)만 삭제됨
        
        # 4. df760 매칭 (Repair_df_sort의 S 1200 열 기준)
        df760_matched = pd.DataFrame()
        if 'df760' in locals() and df760 is not None and not df760.empty:
            # S 1200 열과 df760의 첫 번째 열(index 0)을 기준으로 병합하여 순서 일치시킴
            df760_matched = pd.merge(
                Repair_df_sort[['S 1200']], 
                df760, 
                left_on='S 1200', 
                right_on=df760.columns[0], 
                how='left'
            ).drop(columns=['S 1200']) # 조인에 사용된 중복 키 삭제
        
        # --- 필터 모드일 경우 컬럼 선별 ---
        sort_type = sort_type_var.get()
        if sort_type == "filter":
            # 1. Repair 영역 필터링 (중간 빈 열 "" 포함)
            rep_targets = ["P 1300", "DUT", "PAD", "S/L", "S DUT", "S 정/역", "S PAD", "S S/L", "X 좌표", "Y 좌표"]
            filter_rep = [c for c in Repair_df_sort.columns if str(c).strip() in rep_targets or str(c).strip() == ""]
            Repair_df_sort = Repair_df_sort[filter_rep]
            
            # 2. WWX DATA 영역 필터링
            if df_CM is not None and not df_CM_sort.empty:
                CM_targets = ["STC NO.", "Pin NO."]
                filter_CM = [c for c in df_CM_sort.columns if str(c).strip() in CM_targets]
                if filter_CM:
                    df_CM_sort = df_CM_sort[filter_CM]
            
            # 3. WWX Fail 영역 필터링 (df_fail_matched)
            if not df_fail_matched.empty:
                fail_targets = ["Pad Name", "Coordinate", "X Position", "Y Position", "Error", "Edge Distance", "ERROR", "Scrub X Size", "Scrub Y Size"]
                filter_fail = [c for c in df_fail_matched.columns if str(c).strip() in fail_targets]
                if filter_fail:
                    df_fail_matched = df_fail_matched[filter_fail]
                    
            # 4. 760 Data 영역 필터링
            if not df760_matched.empty:
                t760 = ["No.", "X - 좌표값", "X - 측정값", "X - 오차값", "X - 변환값", 
                        "Y - 좌표값", "Y - 측정값", "Y - 오차값", "Y - 변환값", 
                        "Z - 측정값", "Z - 오차값", "Z - 변환값"]
                filter_760 = [c for c in df760_matched.columns if any(t in str(c) for t in t760)]
                if filter_760:
                    df760_matched = df760_matched[filter_760]

        # 4. 최종 리스트 합치기
        # 행 수는 가장 많은 쪽에 맞춤
        max_rows = max(len(Repair_df_sort), len(df_CM_sort), len(df_fail_matched), len(df760_matched))
        blank = pd.DataFrame({"": [""] * max_rows})
        
        # 기본: Repair_df_sort + 빈칸
        to_concat = [Repair_df_sort.reset_index(drop=True), blank]
        
        # df_CM_sort가 유효하면 추가
        if df_CM is not None and not df_CM_sort.empty:
            to_concat.extend([df_CM_sort.reset_index(drop=True), blank])
            
            # df_fail_matched 추가 (WWX Fail 구역)
            if not df_fail_matched.empty:
                to_concat.extend([df_fail_matched.reset_index(drop=True), blank])
        
        # df760 추가
        to_concat.append(df760_matched.reset_index(drop=True))
        
        final_df = pd.concat(to_concat, axis=1)
        
        # Matching.xlsx 저장
        save_path = os.path.join(os.path.dirname(rep1200_path), "Matching.xlsx")
        
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # 진한 회색 채우기 스타일 설정
            dark_grey_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
            
            # 빈 열(컬럼명이 "")인 위치를 찾아 스타일 적용
            for i, col_name in enumerate(final_df.columns):
                if str(col_name).strip() == "":
                    col_idx = i + 1 # openpyxl은 1부터 시작
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    
                    # 열 너비 설정 (15픽셀은 대략 2.14 캐릭터 너비)
                    worksheet.column_dimensions[col_letter].width = 2.14
                    
                    # 해당 열의 모든 셀에 배경색 적용 (헤더 포함)
                    for row in range(1, max_rows + 2):
                        worksheet.cell(row=row, column=col_idx).fill = dark_grey_fill
                        
        print(f"최종 정렬 및 병합 저장 완료 (스타일 적용): {save_path}")
        check_stop()
        
        # 엑셀 자동 맞춤 (최고 정확도 - COM 방식 활용)
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # 파일 경로를 절대 경로로 변환 (COM 필수)
            abs_save_path = os.path.abspath(save_path)
            check_stop()
            wb = excel.Workbooks.Open(abs_save_path)
            check_stop()
            ws = wb.Worksheets(1)
            check_stop()
            
            # 자동 너비 조정을 원하는 컬럼 목록 (공백 제거 후 비교)
            auto_fit_targets = ["S/L", "S S/L", "S.L", "Coordinate"]
            last_col = ws.UsedRange.Columns.Count
            
            # 첫 번째 행(헤더)을 순회하며 대상 컬럼 찾기
            for c in range(1, last_col + 1):
                header_val = str(ws.Cells(1, c).Value).strip()
                if header_val in auto_fit_targets:
                    ws.Columns(c).AutoFit()
                    print(f"[{header_val}] 컬럼 자동 너비 조정 완료 (COM)")
            
            wb.Save()
            wb.Close()
            excel.Quit()
        except Exception as com_e:
            print(f"Excel AutoFit (COM) 중 오류 발생: {com_e}")
        finally:
            pythoncom.CoUninitialize()
        
        # 미리보기용 데이터 업데이트
        Repair_df = final_df
    except Exception as e:
        if isinstance(e, InterruptedError):
            print("작업이 중단되었습니다.")
            raise
        print("정렬 및 저장 중 오류 발생:", e)

    # 확인을 위해 각 소스 데이터의 크기(행, 열) 및 미리보기 출력
    print("Repair1300 shape (rows, cols):", df1300.shape)
    print("Repair1300 preview (3 rows):\n", df1300.head(3))
    
    print("Repair1200 shape (rows, cols):", df1200.shape)
    print("Repair1200 preview (3 rows):\n", df1200.head(3))
    
    if df_CM is not None:
        print("WWX DATA shape (rows, cols):", df_CM.shape)
        print("WWX DATA preview (3 rows):\n", df_CM.head(3))
        if df_fail is not None:
            print("WWX Fail (2nd Sheet) shape (rows, cols):", df_fail.shape)
            print("WWX Fail preview (3 rows):\n", df_fail.head(3))
    else:
        print("WWX DATA shape: None")
        
    if 'df760' in locals():
        print("760 Data shape (rows, cols):", df760.shape)
        print("760 Data preview (3 rows):\n", df760.head(3))
    else:
        print("760 Data shape: None")
        
    # 최종 리페어 데이터프레임 정보 출력
    print("Repair total rows:", Repair_df.shape[0], "columns:", Repair_df.shape[1])
    print("Repair DataFrame preview (first 3 rows):\n", Repair_df.head(3))
    messagebox.showinfo("완료", "파일 처리가 완료되었습니다. 폴더를 확인해 주세요.")

def create_ui():
    root = tk.Tk()
    root.title("Bonding & waferWoRx")
    root.geometry("800x320") # 높이 확장 (파일명 표시 공간 확보)
    root.configure(bg="#1f1f1f") # 윈도우 다크모드 배경색
    
    # 윈도우 상단 타이틀바 다크모드 적용
    root.update_idletasks()
    try:
        import ctypes
        hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
        # Windows 11 & 최신 Windows 10 (20)
        value = ctypes.c_int(2)
        ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, ctypes.byref(value), ctypes.sizeof(value))
        # 구버전 Windows 10 (19)
        value = ctypes.c_int(1)
        ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 19, ctypes.byref(value), ctypes.sizeof(value))
    except Exception:
        pass
    
    # 스타일 설정
    label_style = {"bg": "#1f1f1f", "fg": "#ffffff", "font": ("Malgun Gothic", 10)}
    entry_style = {"bg": "#2d2d2d", "fg": "#ffffff", "insertbackground": "#ffffff", 
                   "font": ("Malgun Gothic", 10), "width": 90, "relief": "flat"} # 입력창 너비 확장
    # 사진 속 입체감 있는 버튼 스타일 (찾기)
    button_muted = {"bg": "#2d2d2d", "fg": "#ffffff", "font": ("Malgun Gothic", 10), 
                    "width": 8, "relief": "raised", "bd": 2, "activebackground": "#d35400"}
    # 사진 속 입체감 있는 버튼 스타일 (실행)
    button_exec = {"bg": "#2d2d2d", "fg": "#ffffff", "font": ("Malgun Gothic", 10, "bold"), 
                    "width": 10, "relief": "raised", "bd": 2, "activebackground": "#d35400"}
    # 파일명 표시 라벨 스타일
    filename_style = {"bg": "#1f1f1f", "fg": "#888888", "font": ("Malgun Gothic", 9)}
    # 선택 시 주황색으로 변하는 옵션 버튼 스타일
    option_btn_style = {
        "bg": "#2d2d2d", "fg": "#ffffff", "selectcolor": "#d35400", 
        "activebackground": "#d35400", "activeforeground": "#ffffff",
        "font": ("Malgun Gothic", 10), "indicatoron": 0, "width": 8, "relief": "flat"
    }

    fields = [
        ("Repair 1300:", [".xlsx"]),
        ("Repair 1200:", [".xlsx"]),
        ("WWX DATA:", [".xlsm", ".xlsx", ".xls"]),
        ("760 Data:", [".xlsm", ".xlsx", ".xls"]),
    ]
    entries = []
    other_btns = []
    for idx, (label_text, exts) in enumerate(fields):
        # 엔트리와 파일명 라벨을 먼저 생성하여 버튼 명령에서 참조할 수 있게 함
        path_frame = tk.Frame(root, bg="#1f1f1f")
        path_frame.grid(row=idx, column=1, padx=5, pady=10, sticky="w")
        
        ent = tk.Entry(path_frame, **entry_style)
        ent.pack(side="top", anchor="w")
        
        fname_lbl = tk.Label(path_frame, text="", **filename_style)
        fname_lbl.pack(side="top", anchor="w", pady=(2, 0))

        # 라벨 텍스트가 버튼 역할을 하도록 변경 (찾기 버튼 대체)
        lbl_btn = tk.Button(root, text=label_text, 
                            command=lambda e=ent, fl=fname_lbl, ex=exts: select_file(e, fl, ex),
                            bg="#1f1f1f", fg="#ffffff", font=("Malgun Gothic", 10, "bold"),
                            relief="flat", activebackground="#1f1f1f", activeforeground="#ff8c00", 
                            cursor="hand2", bd=0, highlightthickness=0)
        lbl_btn.grid(row=idx, column=0, padx=15, pady=(10, 0), sticky="ne")
        
        entries.append(ent)
        other_btns.append(lbl_btn)

    # 하단 옵션 영역 (방식 및 정렬)
    option_frame = tk.Frame(root, bg="#1f1f1f")
    option_frame.grid(row=len(fields), column=0, columnspan=2, padx=15, pady=10, sticky="w")

    # 정렬 선택 (버튼 스타일)
    tk.Label(option_frame, text="정렬:", **label_style).pack(side="left", padx=(10, 10))
    sort_type_var = tk.StringVar(value="original")
    rb1 = tk.Radiobutton(option_frame, text="원본", variable=sort_type_var, value="original", **option_btn_style)
    rb1.pack(side="left", padx=2)
    rb2 = tk.Radiobutton(option_frame, text="필터", variable=sort_type_var, value="filter", **option_btn_style)
    rb2.pack(side="left", padx=2)
    other_btns.extend([rb1, rb2])

    # 실행 버튼 (쓰레딩 적용하여 응답 없음 방지)
    stop_event = threading.Event()
    is_running = False

    def toggle_task():
        nonlocal is_running
        if not is_running:
            # 실행 시작
            is_running = True
            stop_event.clear()
            exec_btn.config(text="정지", bg="#d35400", activebackground="#ff8c00", state="normal") # 주황색 계열 (정지)
            for b in other_btns:
                b.config(state="disabled")
            
            def run():
                nonlocal is_running
                print("--- 작업 시작 ---")
                pythoncom.CoInitialize()
                try:
                    process_files(entries, sort_type_var, stop_event)
                    print("--- 작업 완료 ---")
                except InterruptedError:
                    print("--- 작업 중단됨 (사용자 요청) ---")
                except Exception as e:
                    print(f"--- 작업 중 오류 발생: {e} ---")
                    root.after(0, lambda: messagebox.showerror("오류", f"작업 중 오류가 발생했습니다:\n{e}"))
                finally:
                    pythoncom.CoUninitialize()
                    is_running = False
                    # UI 업데이트는 메인 쓰레드에서 실행
                    def reset_ui():
                        exec_btn.config(text="실행", state="normal", **button_exec)
                        for b in other_btns:
                            b.config(state="normal")
                        print("--- UI 복구 완료 ---")
                    root.after(0, reset_ui)
            
            threading.Thread(target=run, daemon=True).start()
        else:
            # 정지 요청
            print("--- 중단 요청 중... ---")
            stop_event.set()
            exec_btn.config(state="disabled", text="중단 중...") # 실제 중단될 때까지 비활성화

    exec_btn = tk.Button(root, text="실행", command=toggle_task, **button_exec)
    exec_btn.grid(row=len(fields), column=1, padx=5, pady=10, sticky="e")

    root.mainloop()

if __name__ == "__main__":
    create_ui()
